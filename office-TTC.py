import streamlit as st
import pandas as pd
import json
import time
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from supabase import create_client
import cloudinary
import cloudinary.uploader

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ระบบเบิกอุปกรณ์สำนักงาน",
    page_icon="🏢",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ─── Secrets ──────────────────────────────────────────────────────────────────
# [supabase]
# url = "https://xxxxx.supabase.co"
# key = "eyJhbG..."
# [cloudinary]
# cloud_name = "your_cloud_name"
# api_key = "123456789"
# api_secret = "abcdef..."
# [app]
# admin_password = "admin1234"
# logo_url = ""
# ──────────────────────────────────────────────────────────────────────────────

SUPABASE_URL   = st.secrets["supabase"]["url"]
SUPABASE_KEY   = st.secrets["supabase"]["key"]
ADMIN_PASSWORD = st.secrets["app"]["admin_password"]
LOGO_URL       = st.secrets.get("app", {}).get("logo_url", "")

cloudinary.config(
    cloud_name=st.secrets["cloudinary"]["cloud_name"],
    api_key=st.secrets["cloudinary"]["api_key"],
    api_secret=st.secrets["cloudinary"]["api_secret"],
    secure=True
)

MAX_UPLOAD_MB = 5

# ─── หมวดหมู่อุปกรณ์ ──────────────────────────────────────────────────────────
# group_type: "consumable" = ใช้แล้วหมดไป (ไม่ต้องคืน), "borrow" = ต้องคืน
SUPPLY_GROUPS = {
    "อุปกรณ์สำนักงาน": {
        "icon": "🖊️",
        "type": "consumable",
        "color": "#2D6A4F",
        "desc": "อุปกรณ์ที่ใช้แล้วหมดไป ไม่ต้องคืน",
        "examples": "ปากกา กระดาษ ถ่านไฟฉาย คลิปหนีบกระดาษ น้ำยาลบคำผิด ซองจดหมาย"
    },
    "อุปกรณ์เบิก-คืน": {
        "icon": "🔌",
        "type": "borrow",
        "color": "#D62828",
        "desc": "อุปกรณ์ที่ต้องนำมาคืนหลังใช้งาน",
        "examples": "Dictionary สายต่อ Projector ไมโครโฟน ปากกาไวท์บอร์ด"
    },
    "อุปกรณ์ทั่วไป": {
        "icon": "🧴",
        "type": "consumable",
        "color": "#52796F",
        "desc": "อุปกรณ์สิ้นเปลืองทั่วไป ไม่ต้องคืน",
        "examples": "กระดาษชำระ น้ำยาล้างจาน สบู่ล้างมือ"
    },
}

# ─── Supabase Client ──────────────────────────────────────────────────────────
@st.cache_resource
def get_supabase():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

sb = get_supabase()

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Sarabun', sans-serif;
    font-size: 16px;
    background-color: #f4f6f4;
}
div.stButton > button {
    height: 2.8rem; font-size: 0.95rem; border-radius: 8px;
    font-family: 'Sarabun', sans-serif;
}
input, textarea, select { font-size: 16px !important; }
div[data-baseweb="select"] { font-size: 16px; }

div[data-testid="metric-container"] {
    background: #eaf2ee; border-radius: 10px;
    padding: 12px; border: 1px solid #b7d5c7;
}
.page-title {
    font-size: 1.4rem; font-weight: 700; color: #1b4332;
    margin: 4px 0 16px 0; padding-bottom: 8px;
    border-bottom: 3px solid #2D6A4F;
}
.section-header {
    font-size: 1rem; font-weight: 700; color: #1b4332;
    margin: 16px 0 8px 0; padding: 6px 10px;
    background: #eaf2ee; border-left: 4px solid #2D6A4F;
    border-radius: 0 6px 6px 0;
}
.item-card {
    background: white; border-radius: 10px; padding: 12px 14px;
    margin-bottom: 10px; border: 1px solid #d8e8e0;
    box-shadow: 0 2px 6px rgba(45,106,79,0.06);
}
.group-badge {
    display: inline-flex; align-items: center; gap: 4px;
    padding: 3px 10px; border-radius: 20px;
    font-size: 0.78rem; font-weight: 600; color: white;
}
.overdue-card {
    background: #fff5f5; border-left: 4px solid #D62828;
    border-radius: 0 8px 8px 0; padding: 10px 14px; margin: 6px 0;
}
.consumable-card {
    background: #eaf2ee; border-left: 4px solid #2D6A4F;
    border-radius: 0 8px 8px 0; padding: 10px 14px; margin: 6px 0;
}
.info-box {
    background: #eaf2ee; border: 1px solid #b7d5c7;
    border-radius: 8px; padding: 10px 14px; margin: 8px 0;
    font-size: 0.88rem; color: #1b4332;
}
.qty-badge {
    display: inline-block; padding: 2px 8px; border-radius: 12px;
    font-size: 0.82rem; font-weight: 700;
}
button[kind="primary"] { font-weight: 700; }
</style>
""", unsafe_allow_html=True)

# ─── ADMIN AUTH ───────────────────────────────────────────────────────────────
def is_admin():
    return st.session_state.get("is_admin", False)

def admin_login_widget():
    if is_admin():
        st.sidebar.success("🔓 Admin Mode เปิดอยู่")
        if st.sidebar.button("🔒 ออกจากระบบ Admin", use_container_width=True):
            st.session_state.is_admin = False
            st.rerun()
    else:
        with st.sidebar.expander("🔒 Admin Login"):
            pwd = st.text_input("รหัสผ่าน", type="password", key="admin_pwd")
            if st.button("เข้าสู่ระบบ", use_container_width=True, key="admin_login_btn"):
                if pwd == ADMIN_PASSWORD:
                    st.session_state.is_admin = True
                    st.rerun()
                else:
                    st.error("รหัสผ่านไม่ถูกต้อง")

# ═════════════════════════════════════════════════════════════════════════════
# DATABASE LAYER
# ═════════════════════════════════════════════════════════════════════════════
def _sb_retry(func, retries=2, delay=1):
    for attempt in range(retries + 1):
        try:
            return func()
        except Exception as e:
            if attempt == retries:
                raise e
            time.sleep(delay)

def query_table(table, select="*", filters=None, order=None, limit=None):
    def _do():
        q = sb.table(table).select(select)
        if filters:
            for col, op, val in filters:
                if op == "in_":
                    q = q.in_(col, val)
                else:
                    q = getattr(q, op if op != "is" else "is_")(col, val)
        if order:
            for col_name, opts in order:
                q = q.order(col_name, **opts)
        if limit:
            q = q.limit(limit)
        resp = q.execute()
        return pd.DataFrame(resp.data) if resp.data else pd.DataFrame()
    return _sb_retry(_do)

def insert_row(table, data):
    def _do():
        resp = sb.table(table).insert(data).execute()
        return resp.data[0] if resp.data else None
    return _sb_retry(_do)

def update_rows(table, data, match_col, match_val):
    def _do():
        resp = sb.table(table).update(data).eq(match_col, match_val).execute()
        return resp.data
    return _sb_retry(_do)

def delete_rows(table, match_col=None, match_val=None, delete_all=False):
    def _do():
        if delete_all:
            resp = sb.table(table).delete().gt("id", 0).execute()
        else:
            resp = sb.table(table).delete().eq(match_col, match_val).execute()
        return resp.data
    return _sb_retry(_do)

# ─── Cached Stats ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=30, show_spinner=False)
def load_sidebar_stats():
    df_sup = query_table("supplies", select="id,available_qty,group_type")
    total  = len(df_sup)
    avail  = int(df_sup["available_qty"].sum()) if not df_sup.empty else 0

    today_str = str(date.today())
    df_active = query_table("borrow_transactions", select="id,due_date",
                            filters=[("status","eq","ยืมอยู่")])
    n_borrow = len(df_active)
    n_overdue = len(df_active[df_active["due_date"] < today_str]) if not df_active.empty else 0

    df_consume = query_table("consume_transactions", select="id",
                             limit=1)  # just check exist
    return total, avail, n_borrow, n_overdue

# ─── Batch Enrich ─────────────────────────────────────────────────────────────
def _enrich_borrow(df_tx):
    if df_tx.empty:
        return df_tx
    # cast เป็น int ป้องกัน float key ทำให้ merge ไม่ match
    df_tx = df_tx.copy()
    df_tx["supply_id"]   = df_tx["supply_id"].astype(int)
    df_tx["borrower_id"] = df_tx["borrower_id"].astype(int)
    sup_ids  = df_tx["supply_id"].unique().tolist()
    borr_ids = df_tx["borrower_id"].unique().tolist()

    df_sup  = query_table("supplies", select="id,code,name,group_name,image_url",
                          filters=[("id","in_",sup_ids)]) if sup_ids else pd.DataFrame()
    df_borr = query_table("office_borrowers", select="id,name,type,phone,student_id,department",
                          filters=[("id","in_",borr_ids)]) if borr_ids else pd.DataFrame()

    if not df_sup.empty:
        df_sup["id"] = df_sup["id"].astype(int)
    if not df_borr.empty:
        df_borr["id"] = df_borr["id"].astype(int)

    merged = df_tx.merge(
        df_sup.rename(columns={"id":"sup_id","name":"sup_name","code":"sup_code",
                                "group_name":"sup_group","image_url":"sup_img"}),
        left_on="supply_id", right_on="sup_id", how="left"
    ).merge(
        df_borr.rename(columns={"id":"br_id","name":"br_name","type":"br_type",
                                 "phone":"br_phone","student_id":"br_sid","department":"br_dept"}),
        left_on="borrower_id", right_on="br_id", how="left"
    )
    return merged

def load_active_borrows():
    df = query_table("borrow_transactions",
                     select="id,supply_id,borrower_id,qty,borrow_date,due_date,note,return_date,condition_in,status",
                     filters=[("status","eq","ยืมอยู่")],
                     order=[("due_date",{"desc":False})])
    return _enrich_borrow(df)

def load_pending_borrows():
    df = query_table("borrow_transactions",
                     select="id,supply_id,borrower_id,qty,borrow_date,due_date,return_date,condition_in,note,status",
                     filters=[("status","eq","รอตรวจสอบ")],
                     order=[("return_date",{"desc":False})])
    return _enrich_borrow(df)

# ─── Cloudinary ───────────────────────────────────────────────────────────────
def upload_image(file_obj, public_id):
    file_size_mb = len(file_obj.getvalue()) / (1024 * 1024)
    if file_size_mb > MAX_UPLOAD_MB:
        st.error(f"❌ ไฟล์ใหญ่เกิน {MAX_UPLOAD_MB} MB")
        return None
    try:
        result = cloudinary.uploader.upload(
            file_obj, public_id=public_id, overwrite=True,
            folder="office_supply", resource_type="image",
            transformation=[
                {"width": 1200, "crop": "limit"},
                {"quality": "auto", "fetch_format": "auto"}
            ]
        )
        return result.get("secure_url")
    except Exception as e:
        st.error(f"❌ อัพโหลดรูปไม่สำเร็จ: {e}")
        return None

def optimized_url(original_url, width=400, height=300, crop="pad"):
    if not original_url or not isinstance(original_url, str) or "cloudinary" not in original_url:
        return original_url
    return original_url.replace("/upload/", f"/upload/w_{width},h_{height},c_{crop},q_auto,f_auto/")

def show_image(image_url, width="100%", size="preview"):
    SIZES = {"thumb": (200,150), "preview": (400,300), "full": (800,600)}
    w_img, h_img = SIZES.get(size, (400,300))
    opt = optimized_url(image_url, w_img, h_img, "pad")
    w_style = f"{width}px" if isinstance(width, int) else width
    if opt and isinstance(opt, str) and opt.startswith("http"):
        st.markdown(
            f'<img src="{opt}" style="width:{w_style};max-width:100%;border-radius:8px;'
            f'border:1px solid #ddd;object-fit:contain;" loading="lazy">',
            unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div style="width:{w_style};height:100px;background:#f0f2f6;border-radius:8px;'
            f'display:flex;align-items:center;justify-content:center;font-size:2rem;'
            f'border:1px dashed #ccc;">📦</div>', unsafe_allow_html=True)

# ─── Helpers ──────────────────────────────────────────────────────────────────
def overdue_days(due_str):
    try:
        d = datetime.strptime(str(due_str), "%Y-%m-%d").date()
        return max((date.today() - d).days, 0)
    except (ValueError, TypeError):
        return 0

def group_badge(group_name):
    g = SUPPLY_GROUPS.get(group_name, {})
    icon  = g.get("icon", "📦")
    color = g.get("color", "#888")
    return f'<span class="group-badge" style="background:{color};">{icon} {group_name}</span>'

def is_consumable(group_name):
    return SUPPLY_GROUPS.get(group_name, {}).get("type", "consumable") == "consumable"

# ═════════════════════════════════════════════════════════════════════════════
# NAVIGATION
# ═════════════════════════════════════════════════════════════════════════════
def nav():
    if "page" not in st.session_state:
        st.session_state.page = "Dashboard"

    with st.sidebar:
        st.markdown("## 🏢 ระบบเบิกอุปกรณ์สำนักงาน")
        st.markdown("---")
        admin_login_widget()
        st.markdown("---")
        try:
            total, avail, n_borrow, n_overdue = load_sidebar_stats()
            st.metric("📦 รายการอุปกรณ์", total)
            col1, col2 = st.columns(2)
            col1.metric("✅ คงเหลือ", avail)
            col2.metric("🔄 ยืมอยู่", n_borrow)
            if n_overdue > 0:
                st.error(f"⚠️ เกินกำหนด {n_overdue} รายการ!")
        except Exception:
            st.caption("⏳ กำลังโหลด...")

        st.markdown("---")
        st.markdown("**📂 กลุ่มอุปกรณ์:**")
        for gname, ginfo in SUPPLY_GROUPS.items():
            icon  = ginfo["icon"]
            color = ginfo["color"]
            gtype = "ใช้แล้วหมด" if ginfo["type"] == "consumable" else "เบิก-คืน"
            st.markdown(
                f'<div style="padding:4px 8px;margin:2px 0;border-left:3px solid {color};'
                f'font-size:0.83rem;">{icon} <b>{gname}</b><br>'
                f'<span style="color:#52796F;font-size:0.75rem;">{gtype}</span></div>',
                unsafe_allow_html=True)

    # Logo & header
    logo_html = ""
    if LOGO_URL:
        logo_opt = optimized_url(LOGO_URL, 140, 140, "fit") if "cloudinary" in LOGO_URL else LOGO_URL
        logo_html = f'<img src="{logo_opt}" style="width:60px;height:auto;margin-bottom:4px;" loading="lazy"><br>'

    st.markdown(
        f'<div style="text-align:center;margin-bottom:12px;">'
        f'{logo_html}'
        f'<span style="font-size:1.1rem;font-weight:700;color:#1b4332;">'
        f'🏢 ระบบเบิกอุปกรณ์สำนักงาน</span><br>'
        f'<span style="font-size:0.8rem;color:#52796F;">ภาควิชาครุศาสตร์โยธา — มจพ.</span></div>',
        unsafe_allow_html=True
    )

    NAV = [
        ("Dashboard","🏠","หน้าหลัก"),
        ("คลังอุปกรณ์","📦","คลัง"),
        ("เบิกอุปกรณ์","📋","เบิก"),
        ("คืนอุปกรณ์","↩️","คืน"),
        ("รายงาน","📊","รายงาน"),
        ("ตั้งค่า","⚙️","ตั้งค่า"),
    ]
    cols = st.columns(len(NAV))
    for i, (name, icon, label) in enumerate(NAV):
        active = st.session_state.page == name
        with cols[i]:
            if st.button(f"{icon}\n{label}", key=f"nav_{name}",
                         type="primary" if active else "secondary",
                         use_container_width=True):
                st.session_state.page = name
                st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
def page_dashboard():
    st.markdown('<div class="page-title">🏠 ภาพรวมระบบ</div>', unsafe_allow_html=True)

    try:
        total, avail, n_borrow, n_overdue = load_sidebar_stats()
    except Exception:
        st.error("❌ ไม่สามารถเชื่อมต่อ Supabase กรุณาตรวจสอบการตั้งค่า")
        return

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 รายการทั้งหมด", total)
    c2.metric("✅ หน่วยคงเหลือ", avail)
    c3.metric("🔄 กำลังยืม", n_borrow)
    c4.metric("⚠️ เกินกำหนด", n_overdue, delta=f"{n_overdue}" if n_overdue > 0 else None,
              delta_color="inverse" if n_overdue > 0 else "off")

    # กลุ่มอุปกรณ์ summary
    st.markdown('<div class="section-header">📂 สรุปตามกลุ่ม</div>', unsafe_allow_html=True)
    df_sup = query_table("supplies", select="id,available_qty,total_qty,group_name")
    if not df_sup.empty:
        cols = st.columns(3)
        for i, (gname, ginfo) in enumerate(SUPPLY_GROUPS.items()):
            df_g = df_sup[df_sup["group_name"] == gname] if not df_sup.empty else pd.DataFrame()
            n_items = len(df_g)
            n_avail = int(df_g["available_qty"].sum()) if not df_g.empty else 0
            with cols[i]:
                color = ginfo["color"]
                icon  = ginfo["icon"]
                st.markdown(
                    f'<div class="item-card" style="border-top:4px solid {color};text-align:center;">'
                    f'<div style="font-size:2rem;margin-bottom:4px;">{icon}</div>'
                    f'<div style="font-weight:700;color:{color};font-size:0.95rem;">{gname}</div>'
                    f'<div style="font-size:0.8rem;color:#52796F;margin:2px 0;">{ginfo["desc"]}</div>'
                    f'<hr style="margin:8px 0;border-color:#eee;">'
                    f'<div style="font-size:1.1rem;font-weight:700;">{n_items} รายการ</div>'
                    f'<div style="font-size:0.85rem;color:#555;">คงเหลือ {n_avail} หน่วย</div>'
                    f'</div>', unsafe_allow_html=True)

    # รายการยืมเกินกำหนด
    if n_overdue > 0:
        st.markdown('<div class="section-header" style="border-color:#D62828;background:#fff5f5;">⚠️ เกินกำหนดคืน</div>', unsafe_allow_html=True)
        df_active = load_active_borrows()
        if not df_active.empty:
            today_str = str(date.today())
            df_od = df_active[df_active["due_date"] < today_str].copy()
            df_od["days_over"] = df_od["due_date"].apply(overdue_days)
            df_od = df_od.sort_values("days_over", ascending=False)
            for _, r in df_od.iterrows():
                phone = f" | 📞 {r['br_phone']}" if pd.notna(r.get("br_phone")) and r.get("br_phone") else ""
                st.markdown(
                    f'<div class="overdue-card">'
                    f'🔴 <b>{r["sup_code"]} — {r["sup_name"]}</b> ({r["qty"]} ชิ้น)<br>'
                    f'👤 {r["br_name"]} ({r["br_type"]}){phone}<br>'
                    f'📅 กำหนดคืน: {r["due_date"]} '
                    f'<b style="color:#D62828;">เกิน {r["days_over"]} วัน</b>'
                    f'</div>', unsafe_allow_html=True)

    # รายการยืมอยู่
    st.markdown('<div class="section-header">🔄 รายการที่กำลังยืมอยู่</div>', unsafe_allow_html=True)
    df_active = load_active_borrows()
    if df_active.empty:
        st.info("✅ ไม่มีรายการยืมในขณะนี้")
    else:
        today_str = str(date.today())
        for _, r in df_active.iterrows():
            od = overdue_days(r["due_date"])
            bc = "#D62828" if od > 0 else "#2D6A4F"
            st.markdown(
                f'<div class="item-card" style="border-left:4px solid {bc};">'
                f'{group_badge(r.get("sup_group",""))}'
                f' <b>{r["sup_code"]}</b> — {r["sup_name"]} '
                f'<span class="qty-badge" style="background:#e0eeea;color:#1b4332;">{r["qty"]} ชิ้น</span><br>'
                f'👤 {r["br_name"]} ({r["br_type"]})<br>'
                f'📅 เบิก {r["borrow_date"]} | กำหนดคืน <b>{r["due_date"]}</b>'
                + (f' &nbsp;<b style="color:#D62828;">⚠️ เกิน {od} วัน</b>' if od > 0 else "")
                + '</div>', unsafe_allow_html=True)

    # ประวัติเบิกล่าสุด (consumable)
    st.markdown('<div class="section-header">📋 การเบิกอุปกรณ์สิ้นเปลืองล่าสุด</div>', unsafe_allow_html=True)
    df_con = query_table("consume_transactions",
                         select="id,supply_id,requester_name,qty,request_date,purpose",
                         order=[("id",{"desc":True})], limit=5)
    if df_con.empty:
        st.info("ยังไม่มีประวัติการเบิก")
    else:
        sup_ids = df_con["supply_id"].unique().tolist()
        df_s = query_table("supplies", select="id,code,name,group_name",
                           filters=[("id","in_",sup_ids)])
        merged = df_con.merge(
            df_s.rename(columns={"id":"sid","name":"sname","code":"scode","group_name":"sgroup"}),
            left_on="supply_id", right_on="sid", how="left"
        )
        for _, r in merged.iterrows():
            st.markdown(
                f'<div class="consumable-card">'
                f'{group_badge(r.get("sgroup",""))} '
                f'<b>{r["scode"]}</b> — {r["sname"]} '
                f'<span class="qty-badge" style="background:#d8eedf;color:#1b4332;">{r["qty"]} ชิ้น</span><br>'
                f'👤 {r["requester_name"]} | 📅 {r["request_date"]}'
                + (f' | 📝 {r["purpose"]}' if r.get("purpose") else "")
                + '</div>', unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: INVENTORY (คลังอุปกรณ์)
# ═════════════════════════════════════════════════════════════════════════════
def page_inventory():
    st.markdown('<div class="page-title">📦 คลังอุปกรณ์สำนักงาน</div>', unsafe_allow_html=True)

    # Filters
    col_s, col_g = st.columns([2, 1])
    with col_s:
        search = st.text_input("🔍 ค้นหา ชื่อ / รหัส", placeholder="พิมพ์เพื่อค้นหา...")
    with col_g:
        group_sel = st.selectbox("📂 กลุ่ม", ["ทั้งหมด"] + list(SUPPLY_GROUPS.keys()), key="inv_group")

    df_all = query_table("supplies",
                         select="id,code,name,group_name,unit,total_qty,available_qty,status,image_url,description,min_qty",
                         order=[("group_name",{"desc":False}),("code",{"desc":False})])

    df = df_all.copy()
    if search:
        mask = (df["name"].str.contains(search, case=False, na=False) |
                df["code"].str.contains(search, case=False, na=False))
        df = df[mask]
    if group_sel != "ทั้งหมด":
        df = df[df["group_name"] == group_sel]

    st.caption(f"พบ {len(df)} รายการ")

    if df.empty:
        st.info("ไม่พบรายการที่ค้นหา")
    else:
        # batch last transaction
        df_last = pd.DataFrame()
        df_borr_all = pd.DataFrame()
        df_all_tx = query_table("borrow_transactions",
                                select="id,supply_id,borrower_id,borrow_date",
                                order=[("id",{"desc":True})])
        if not df_all_tx.empty:
            df_last = df_all_tx.drop_duplicates(subset=["supply_id"], keep="first")
            br_ids = df_last["borrower_id"].unique().tolist()
            if br_ids:
                df_borr_all = query_table("office_borrowers", select="id,name",
                                          filters=[("id","in_",br_ids)])

        for _, r in df.iterrows():
            gname  = r.get("group_name","")
            ginfo  = SUPPLY_GROUPS.get(gname, {"color":"#888","icon":"📦","type":"consumable"})
            color  = ginfo["color"]
            icon   = ginfo["icon"]
            avail  = int(r.get("available_qty", 0))
            total  = int(r.get("total_qty", 0))
            min_q  = int(r.get("min_qty", 0)) if pd.notna(r.get("min_qty")) else 0
            low_stock = min_q > 0 and avail <= min_q
            status_color = "#D62828" if r.get("status") != "พร้อมใช้" else color

            with st.expander(
                f"{icon} {r['code']} — {r['name']}"
                + (" ⚠️ สต็อกต่ำ!" if low_stock else ""),
                expanded=False
            ):
                c1, c2 = st.columns([1, 2])
                with c1:
                    show_image(r.get("image_url"), width="100%", size="preview")
                with c2:
                    st.markdown(group_badge(gname), unsafe_allow_html=True)
                    st.markdown(
                        f"**รหัส:** {r['code']}<br>"
                        f"**หน่วย:** {r.get('unit','ชิ้น')}<br>"
                        f"**คงเหลือ:** <b style='color:{status_color};'>{avail}</b> / {total}<br>"

                        + f"**สถานะ:** {r.get('status','พร้อมใช้')}",
                        unsafe_allow_html=True)
                    if r.get("description"):
                        st.caption(r["description"])
                    if low_stock:
                        st.warning(f"⚠️ สต็อกใกล้หมด! คงเหลือ {avail} {r.get('unit','ชิ้น')} (ขั้นต่ำ {min_q})")

                # ผู้ยืมล่าสุด (เฉพาะ borrow type)
                if not is_consumable(gname) and not df_last.empty:
                    lt = df_last[df_last["supply_id"] == r["id"]]
                    if not lt.empty and not df_borr_all.empty:
                        br = df_borr_all[df_borr_all["id"] == lt.iloc[0]["borrower_id"]]
                        if not br.empty:
                            st.caption(f"👤 ยืมล่าสุด: **{br.iloc[0]['name']}** ({lt.iloc[0]['borrow_date']})")

                # Admin actions
                if is_admin():
                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button("✏️ แก้ไข", key=f"edit_sup_{r['id']}", use_container_width=True):
                            st.session_state["_edit_supply_id"] = int(r["id"])
                            st.session_state.page = "คลังอุปกรณ์"
                            st.rerun()
                    with bc2:
                        if st.button("🗑️ ลบ", key=f"del_sup_{r['id']}", use_container_width=True):
                            # ตรวจสอบว่ามีการยืมอยู่
                            active_check = query_table("borrow_transactions", select="id",
                                                       filters=[("supply_id","eq",r["id"]),
                                                                ("status","eq","ยืมอยู่")])
                            if len(active_check) > 0:
                                st.error("❌ ลบไม่ได้ มีการยืมอยู่")
                            else:
                                delete_rows("supplies", "id", r["id"])
                                load_sidebar_stats.clear()
                                st.success("✅ ลบแล้ว")
                                st.rerun()

    # ── Admin: เพิ่ม/แก้ไข ─────────────────────────────────────────────────
    if is_admin():
        st.markdown('<div class="section-header">➕ เพิ่ม / แก้ไขรายการ</div>', unsafe_allow_html=True)

        options = ["➕ เพิ่มใหม่"] + [f"{r['code']} — {r['name']}" for _, r in df_all.iterrows()]

        # ถ้ากดปุ่ม ✏️ แก้ไข จาก card ให้เซ็ต _next_sup_sel ก่อน
        if "_edit_supply_id" in st.session_state:
            _eid = st.session_state.pop("_edit_supply_id")
            _match = df_all[df_all["id"] == _eid]
            if not _match.empty:
                _lbl = f"{_match.iloc[0]['code']} — {_match.iloc[0]['name']}"
                if _lbl in options:
                    st.session_state["_next_sup_sel"] = _lbl

        # pattern เดียวกับ Lab app — เซ็ต widget key ก่อน render
        if "_next_sup_sel" in st.session_state:
            st.session_state["sup_edit_sel"] = st.session_state.pop("_next_sup_sel")

        choice = st.selectbox("เลือกรายการ", options, key="sup_edit_sel")

        existing = None
        sup_id = None
        if choice != "➕ เพิ่มใหม่":
            code_sel = choice.split(" — ")[0].strip()
            match = df_all[df_all["code"] == code_sel]
            if not match.empty:
                existing = match.iloc[0]
                sup_id = int(existing["id"])

        with st.form("form_supply"):
            col_a, col_b = st.columns(2)
            with col_a:
                sv_code = st.text_input("รหัส *", value=str(existing["code"]) if existing is not None else "")
                sv_name = st.text_input("ชื่ออุปกรณ์ *", value=str(existing["name"]) if existing is not None else "")
                sv_unit = st.text_input("หน่วย", value=str(existing.get("unit","ชิ้น")) if existing is not None else "ชิ้น")
            with col_b:
                sv_group = st.selectbox("กลุ่ม *", list(SUPPLY_GROUPS.keys()),
                    index=list(SUPPLY_GROUPS.keys()).index(existing["group_name"])
                          if existing is not None and existing.get("group_name") in SUPPLY_GROUPS else 0)
                sv_qty = st.number_input("จำนวนทั้งหมด", min_value=0,
                    value=int(existing["total_qty"]) if existing is not None else 1)
                sv_min = st.number_input("จำนวนขั้นต่ำ (แจ้งเตือน)", min_value=0,
                    value=int(existing["min_qty"]) if existing is not None and pd.notna(existing.get("min_qty")) else 0)

            sv_status = st.selectbox("สถานะ", ["พร้อมใช้","หมด","ชำรุด"],
                index=["พร้อมใช้","หมด","ชำรุด"].index(str(existing.get("status","พร้อมใช้")))
                      if existing is not None else 0)
            sv_desc = st.text_area("รายละเอียด/หมายเหตุ",
                value=str(existing.get("description","") or "") if existing is not None else "")
            sv_img = st.file_uploader(f"📷 รูป (สูงสุด {MAX_UPLOAD_MB} MB)", type=["jpg","jpeg","png"])
            if existing is not None and existing.get("image_url"):
                st.caption("รูปปัจจุบัน:")
                show_image(existing["image_url"], size="thumb")

            # แสดงข้อมูล group ที่เลือก
            ginfo_sel = SUPPLY_GROUPS.get(sv_group, {})
            st.markdown(
                f'<div class="info-box">'
                f'{ginfo_sel.get("icon","")} <b>{sv_group}</b>: {ginfo_sel.get("desc","")}<br>'
                f'ตัวอย่าง: {ginfo_sel.get("examples","")}</div>',
                unsafe_allow_html=True)

            submitted = st.form_submit_button("💾 บันทึก", type="primary", use_container_width=True)

        if submitted:
            sv_code = sv_code.strip()
            sv_name = sv_name.strip()
            if not sv_code or not sv_name:
                st.error("❌ กรุณากรอกรหัสและชื่ออุปกรณ์")
            else:
                dup = query_table("supplies", select="id", filters=[("code","eq",sv_code)])
                if not dup.empty and (sup_id is None or int(dup.iloc[0]["id"]) != sup_id):
                    st.error(f"❌ รหัส '{sv_code}' มีอยู่แล้ว")
                else:
                    try:
                        img_url = existing.get("image_url") if existing is not None else None
                        if sv_img:
                            img_url = upload_image(sv_img, f"office_{sv_code}")
                            if img_url is None:
                                st.stop()

                        gtype = SUPPLY_GROUPS.get(sv_group,{}).get("type","consumable")
                        payload = {
                            "code": sv_code, "name": sv_name, "unit": sv_unit or "ชิ้น",
                            "group_name": sv_group, "group_type": gtype,
                            "total_qty": sv_qty,
                            "status": sv_status, "image_url": img_url,
                            "description": sv_desc or None,
                            "min_qty": sv_min if sv_min > 0 else None
                        }

                        if existing is None:
                            payload["available_qty"] = sv_qty
                            insert_row("supplies", payload)
                            st.success(f"✅ เพิ่ม '{sv_code} — {sv_name}' เรียบร้อย")
                            st.session_state["_next_sup_sel"] = "➕ เพิ่มใหม่"
                        else:
                            old_total = int(existing["total_qty"])
                            old_avail = int(existing["available_qty"])
                            diff = sv_qty - old_total
                            new_avail = old_avail + diff
                            if new_avail < 0:
                                st.error(f"❌ ลดจำนวนไม่ได้! ยืมออกอยู่ {old_total - old_avail} ชิ้น")
                                st.stop()
                            payload["available_qty"] = new_avail
                            update_rows("supplies", payload, "id", sup_id)
                            st.success(f"✅ แก้ไข '{sv_code} — {sv_name}' เรียบร้อย")
                            # คงอยู่ที่รายการเดิมหลังบันทึก
                            st.session_state["_next_sup_sel"] = f"{sv_code} — {sv_name}"

                        load_sidebar_stats.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ เกิดข้อผิดพลาด: {e}")
    else:
        st.info("🔒 การเพิ่ม/แก้ไขอุปกรณ์ สำหรับ Admin เท่านั้น")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: BORROW/REQUEST (เบิกอุปกรณ์)
# ═════════════════════════════════════════════════════════════════════════════
def page_request():
    st.markdown('<div class="page-title">📋 เบิกอุปกรณ์</div>', unsafe_allow_html=True)

    # แยก tab ตามประเภท
    tab_con, tab_borrow = st.tabs(["🖊️ เบิกอุปกรณ์สิ้นเปลือง", "🔌 เบิก-ยืมอุปกรณ์คืน"])

    # ── TAB 1: เบิกอุปกรณ์สิ้นเปลือง (consumable) ───────────────────────────
    with tab_con:
        st.markdown(
            '<div class="info-box">🖊️ อุปกรณ์ <b>สำนักงาน</b> และ <b>ทั่วไป</b> — '
            'ใช้แล้วหมดไป <b>ไม่ต้องคืน</b></div>',
            unsafe_allow_html=True)

        df_con_sup = query_table("supplies",
                                 select="id,code,name,group_name,unit,available_qty,image_url,description",
                                 filters=[("group_type","eq","consumable"),("status","eq","พร้อมใช้")],
                                 order=[("group_name",{"desc":False}),("code",{"desc":False})])
        df_con_sup = df_con_sup[df_con_sup["available_qty"] > 0] if not df_con_sup.empty else df_con_sup

        if df_con_sup.empty:
            st.warning("⚠️ ไม่มีอุปกรณ์สิ้นเปลืองพร้อมเบิก")
        else:
            # กรอง group
            groups_avail = sorted(df_con_sup["group_name"].dropna().unique().tolist())
            g_sel = st.selectbox("📂 กลุ่มอุปกรณ์", ["ทั้งหมด"] + groups_avail, key="con_grp")
            search_c = st.text_input("🔍 ค้นหา", placeholder="ชื่อ หรือ รหัส", key="con_search")

            df_f = df_con_sup.copy()
            if g_sel != "ทั้งหมด":
                df_f = df_f[df_f["group_name"] == g_sel]
            if search_c:
                mask = (df_f["name"].str.contains(search_c, case=False, na=False) |
                        df_f["code"].str.contains(search_c, case=False, na=False))
                df_f = df_f[mask]

            if df_f.empty:
                st.info("ไม่พบรายการ")
            else:
                opts = {f"{r['code']} — {r['name']}  ({r['available_qty']} {r.get('unit','ชิ้น')} คงเหลือ)": r["id"]
                        for _, r in df_f.iterrows()}
                sel_lbl = st.selectbox("เลือกอุปกรณ์ *", list(opts.keys()), key="con_eq")
                eq_row = df_f[df_f["id"] == opts[sel_lbl]].iloc[0]

                c_img, c_info = st.columns([1, 2])
                with c_img:
                    show_image(eq_row.get("image_url"), size="preview")
                with c_info:
                    st.markdown(group_badge(eq_row.get("group_name","")), unsafe_allow_html=True)
                    st.markdown(f"**{eq_row['code']}** — {eq_row['name']}")
                    st.markdown(f"คงเหลือ: **{eq_row['available_qty']}** {eq_row.get('unit','ชิ้น')}")
                    if eq_row.get("description"):
                        st.caption(eq_row["description"])

                qty_con = st.number_input("จำนวนที่ต้องการเบิก *", min_value=1,
                                          max_value=int(eq_row["available_qty"]), value=1, key="con_qty")
                st.divider()
                st.markdown('<div class="section-header">👤 ข้อมูลผู้เบิก</div>', unsafe_allow_html=True)
                req_name = st.text_input("ชื่อ-นามสกุล *", placeholder="กรอกชื่อ-นามสกุล", key="con_name")
                req_type = st.radio("ประเภท", ["นักศึกษา","บุคลากร/อาจารย์"], horizontal=True, key="con_type")
                req_dept = st.text_input("ภาควิชา / หน่วยงาน", key="con_dept")
                purpose  = st.text_input("วัตถุประสงค์การเบิก (ถ้ามี)", key="con_purpose")
                req_date = st.date_input("วันที่เบิก", value=date.today(), key="con_date")

                if st.button("✅ ยืนยันการเบิก", type="primary", use_container_width=True, key="con_submit"):
                    if not req_name.strip():
                        st.error("❌ กรุณากรอกชื่อผู้เบิก")
                    elif req_date > date.today():
                        st.error("❌ ไม่อนุญาตให้เบิกล่วงหน้า กรุณาใช้วันที่ปัจจุบันหรือก่อนหน้า")
                    else:
                        try:
                            insert_row("consume_transactions", {
                                "supply_id": int(opts[sel_lbl]),
                                "requester_name": req_name.strip(),
                                "requester_type": req_type,
                                "department": req_dept or None,
                                "qty": qty_con,
                                "request_date": str(req_date),
                                "purpose": purpose or None,
                                "status": "เบิกแล้ว"
                            })
                            new_avail = int(eq_row["available_qty"]) - qty_con
                            update_rows("supplies", {"available_qty": new_avail}, "id", int(opts[sel_lbl]))
                            # อัพเดทสถานะถ้าหมด
                            if new_avail <= 0:
                                update_rows("supplies", {"status": "หมด"}, "id", int(opts[sel_lbl]))
                            load_sidebar_stats.clear()
                            st.success(
                                f"✅ บันทึกสำเร็จ!\n\n"
                                f"👤 **{req_name}** เบิก {eq_row['name']} จำนวน {qty_con} {eq_row.get('unit','ชิ้น')}"
                            )
                            st.balloons()
                        except Exception as e:
                            st.error(f"❌ เกิดข้อผิดพลาด: {e}")

    # ── TAB 2: เบิก-ยืม (borrow type — ต้องคืน) ─────────────────────────────
    with tab_borrow:
        st.markdown(
            '<div class="info-box">🔌 อุปกรณ์ <b>เบิก-คืน</b> — '
            'ต้องนำมาคืนหลังใช้งาน Admin จะตรวจสอบสภาพ</div>',
            unsafe_allow_html=True)

        df_bor_sup = query_table("supplies",
                                 select="id,code,name,group_name,unit,available_qty,image_url,description",
                                 filters=[("group_type","eq","borrow"),("status","eq","พร้อมใช้")],
                                 order=[("code",{"desc":False})])
        df_bor_sup = df_bor_sup[df_bor_sup["available_qty"] > 0] if not df_bor_sup.empty else df_bor_sup

        if df_bor_sup.empty:
            st.warning("⚠️ ไม่มีอุปกรณ์เบิก-คืนพร้อมใช้งาน")
        else:
            search_b = st.text_input("🔍 ค้นหา", placeholder="ชื่อ หรือ รหัส", key="bor_search")
            df_fb = df_bor_sup.copy()
            if search_b:
                mask = (df_fb["name"].str.contains(search_b, case=False, na=False) |
                        df_fb["code"].str.contains(search_b, case=False, na=False))
                df_fb = df_fb[mask]

            if df_fb.empty:
                st.info("ไม่พบรายการ")
            else:
                opts_b = {f"{r['code']} — {r['name']}  ({r['available_qty']} {r.get('unit','ชิ้น')} คงเหลือ)": r["id"]
                          for _, r in df_fb.iterrows()}
                sel_b = st.selectbox("เลือกอุปกรณ์ *", list(opts_b.keys()), key="bor_eq")
                eq_b  = df_fb[df_fb["id"] == opts_b[sel_b]].iloc[0]

                c_img2, c_info2 = st.columns([1, 2])
                with c_img2:
                    show_image(eq_b.get("image_url"), size="preview")
                with c_info2:
                    st.markdown(group_badge(eq_b.get("group_name","")), unsafe_allow_html=True)
                    st.markdown(f"**{eq_b['code']}** — {eq_b['name']}")
                    st.markdown(f"คงเหลือ: **{eq_b['available_qty']}** {eq_b.get('unit','ชิ้น')}")
                    if eq_b.get("description"):
                        st.caption(eq_b["description"])

                qty_b = st.number_input("จำนวน *", min_value=1,
                                        max_value=int(eq_b["available_qty"]), value=1, key="bor_qty")
                st.divider()
                st.markdown('<div class="section-header">👤 ข้อมูลผู้ยืม</div>', unsafe_allow_html=True)
                bor_type = st.radio("ประเภท", ["นักศึกษา","บุคลากร/อาจารย์"], horizontal=True, key="bor_type")
                bor_name = st.text_input("ชื่อ-นามสกุล *", key="bor_name")
                bor_sid  = st.text_input("รหัสนักศึกษา / รหัสพนักงาน", key="bor_sid")
                bor_dept = st.text_input("ภาควิชา / หน่วยงาน", key="bor_dept")
                bor_phone= st.text_input("เบอร์โทรศัพท์ *", key="bor_phone")
                st.divider()
                st.markdown('<div class="section-header">📅 วันที่</div>', unsafe_allow_html=True)
                borrow_date = st.date_input("วันที่เบิก *", value=date.today(), key="bor_date")
                due_date    = st.date_input("วันกำหนดคืน *", value=date.today(), key="bor_due")
                cond_out    = st.selectbox("สภาพอุปกรณ์ขณะเบิก",
                                           ["ปกติ","มีรอยขีดข่วน","ชำรุดบางส่วน"], key="bor_cond")
                note_b = st.text_area("หมายเหตุ", key="bor_note")

                if st.button("✅ ยืนยันการยืม", type="primary", use_container_width=True, key="bor_submit"):
                    if not bor_name.strip():
                        st.error("❌ กรุณากรอกชื่อ")
                    elif not bor_phone.strip():
                        st.error("❌ กรุณากรอกเบอร์โทรศัพท์")
                    elif borrow_date > date.today():
                        st.error("❌ ไม่อนุญาตให้เบิกล่วงหน้า กรุณาใช้วันที่ปัจจุบันหรือก่อนหน้า")
                    elif due_date < borrow_date:
                        st.error("❌ วันกำหนดคืนต้องไม่ก่อนวันที่เบิก")
                    else:
                        try:
                            existing_borr = query_table("office_borrowers", select="id",
                                                        filters=[("phone","eq",bor_phone.strip())])
                            if not existing_borr.empty:
                                borr_id = int(existing_borr.iloc[0]["id"])
                                update_rows("office_borrowers", {
                                    "name": bor_name.strip(), "type": bor_type,
                                    "student_id": bor_sid or None, "department": bor_dept or None
                                }, "id", borr_id)
                            else:
                                borr = insert_row("office_borrowers", {
                                    "name": bor_name.strip(), "type": bor_type,
                                    "student_id": bor_sid or None, "department": bor_dept or None,
                                    "phone": bor_phone.strip()
                                })
                                borr_id = int(borr["id"])

                            insert_row("borrow_transactions", {
                                "supply_id": int(opts_b[sel_b]),
                                "borrower_id": borr_id,
                                "qty": qty_b,
                                "borrow_date": str(borrow_date),
                                "due_date": str(due_date),
                                "condition_out": cond_out,
                                "note": note_b or None,
                                "status": "ยืมอยู่"
                            })
                            new_avail_b = int(eq_b["available_qty"]) - qty_b
                            update_rows("supplies", {"available_qty": new_avail_b}, "id", int(opts_b[sel_b]))
                            if new_avail_b <= 0:
                                update_rows("supplies", {"status": "ยืมออก"}, "id", int(opts_b[sel_b]))
                            load_sidebar_stats.clear()

                            st.success(
                                f"✅ บันทึกสำเร็จ!\n\n"
                                f"👤 **{bor_name}** ยืม {eq_b['name']} x{qty_b}\n"
                                f"📅 กำหนดคืน: {due_date}"
                            )
                            st.balloons()
                        except Exception as e:
                            st.error(f"❌ เกิดข้อผิดพลาด: {e}")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: RETURN (คืนอุปกรณ์)
# ═════════════════════════════════════════════════════════════════════════════
def page_return():
    st.markdown('<div class="page-title">↩️ คืนอุปกรณ์</div>', unsafe_allow_html=True)

    df_pending = load_pending_borrows()
    n_pending = len(df_pending)
    # นับรายการที่ยกเลิกได้ (ยืมอยู่ + รอตรวจสอบ)
    df_cancel_list = query_table("borrow_transactions", select="id",
                                 filters=[("status","in_",["ยืมอยู่","รอตรวจสอบ"])])
    n_cancel = len(df_cancel_list)
    tab1, tab2, tab3 = st.tabs(["📬 แจ้งคืน", f"🔍 รอตรวจสอบ ({n_pending})", f"🚫 ยกเลิกรายการ ({n_cancel})"])

    with tab1:
        st.markdown(
            '<div class="info-box">📬 ผู้ยืมกรอกข้อมูลแจ้งคืน → Admin ตรวจสอบและยืนยัน</div>',
            unsafe_allow_html=True)
        search_r = st.text_input("🔍 ค้นหาชื่อ / รหัสอุปกรณ์", key="ret_search")

        df_active = load_active_borrows()
        if search_r and not df_active.empty:
            s = search_r.lower()
            mask = (df_active["br_name"].str.lower().str.contains(s, na=False) |
                    df_active["sup_code"].str.lower().str.contains(s, na=False) |
                    df_active["sup_name"].str.lower().str.contains(s, na=False))
            df_active = df_active[mask]

        if df_active.empty:
            st.info("✅ ไม่มีรายการยืม" if not search_r else "ไม่พบรายการที่ค้นหา")
        else:
            st.caption(f"พบ {len(df_active)} รายการ")
            for _, r in df_active.iterrows():
                od = overdue_days(r["due_date"])
                lbl = f"TX#{r['id']} | {r['sup_code']} {r['sup_name']} | {r['br_name']}"
                if od > 0:
                    lbl += f" ⚠️ เกิน {od} วัน"
                with st.expander(lbl):
                    c1, c2 = st.columns([1, 2])
                    with c1:
                        show_image(r.get("sup_img"), size="preview")
                    with c2:
                        st.markdown(
                            f"📦 **{r['sup_code']}** — {r['sup_name']} ({r['qty']} ชิ้น)<br>"
                            f"👤 {r['br_name']} ({r['br_type']})<br>"
                            f"📅 เบิก {r['borrow_date']} | คืน <b>{r['due_date']}</b>"
                            + (f"<br><b style='color:#D62828;'>⚠️ เกินกำหนด {od} วัน</b>" if od > 0 else ""),
                            unsafe_allow_html=True)
                        if r.get("br_phone"):
                            st.caption(f"📞 {r['br_phone']}")

                    ret_date = st.date_input("วันที่นำมาคืน", value=date.today(), key=f"rd_{r['id']}")
                    cond_in  = st.selectbox("สภาพอุปกรณ์",
                                            ["ปกติ","มีรอยขีดข่วน","ชำรุด","สูญหาย"],
                                            key=f"ci_{r['id']}")
                    ret_name = st.text_input("ชื่อผู้นำมาคืน *",
                                             value=str(r.get("br_name","")) if pd.notna(r.get("br_name","")) else "",
                                             key=f"rname_{r['id']}")
                    note_r = st.text_input("หมายเหตุ", key=f"rn_{r['id']}")

                    if st.button("📬 แจ้งคืน", key=f"notify_{r['id']}",
                                 type="primary", use_container_width=True):
                        borrow_dt = datetime.strptime(str(r["borrow_date"]), "%Y-%m-%d").date()
                        if not ret_name.strip():
                            st.error("❌ กรุณากรอกชื่อผู้นำมาคืน")
                        elif ret_date < borrow_dt:
                            st.error(f"❌ วันที่คืนต้องไม่ก่อนวันที่เบิก ({r['borrow_date']})")
                        else:
                            update_rows("borrow_transactions", {
                                "return_date": str(ret_date),
                                "condition_in": cond_in,
                                "note": note_r or None,
                                "status": "รอตรวจสอบ"
                            }, "id", r["id"])
                            st.success("📬 แจ้งคืนแล้ว! รอ Admin ตรวจสอบ")
                            load_sidebar_stats.clear()
                            st.rerun()

    with tab2:
        if df_pending.empty:
            st.info("✅ ไม่มีรายการรอตรวจสอบ")
        else:
            st.warning(f"🔍 {len(df_pending)} รายการรอ Admin ตรวจสอบ")
            for _, r in df_pending.iterrows():
                od = overdue_days(r["due_date"])
                lbl = f"TX#{r['id']} | {r['sup_code']} {r['sup_name']} | {r['br_name']}"
                with st.expander(lbl):
                    c1, c2 = st.columns([1, 2])
                    with c1:
                        show_image(r.get("sup_img"), size="preview")
                    with c2:
                        phone_str = f"📞 {r['br_phone']}" if pd.notna(r.get("br_phone")) and r.get("br_phone") else ""
                        st.markdown(
                            f"📦 **{r['sup_code']}** — {r['sup_name']} ({r['qty']} ชิ้น)<br>"
                            f"👤 {r['br_name']} ({r['br_type']}) {phone_str}<br>"
                            f"📅 เบิก {r['borrow_date']} | กำหนดคืน <b>{r['due_date']}</b><br>"
                            f"📅 แจ้งคืนวันที่: <b>{r.get('return_date','')}</b>"
                            + (f"<br><b style='color:#D62828;'>⚠️ เกิน {od} วัน</b>" if od > 0 else ""),
                            unsafe_allow_html=True)
                        st.markdown(f"**สภาพตอนเบิก:** {r.get('condition_out','-')}")
                        st.markdown(f"**สภาพที่แจ้งคืน:** {r.get('condition_in','-')}")
                        if r.get("note"):
                            st.caption(f"หมายเหตุ: {r['note']}")

                    if is_admin():
                        st.markdown("**🔍 Admin ตรวจสอบ:**")
                        admin_cond = st.selectbox("สภาพจริง",
                                                   ["ปกติ","มีรอยขีดข่วน","ชำรุด","สูญหาย"],
                                                   key=f"ac_{r['id']}")
                        admin_note = st.text_input("หมายเหตุ Admin", key=f"an_{r['id']}")
                        cc1, cc2 = st.columns(2)
                        with cc1:
                            if st.button("✅ ยืนยันรับคืน", key=f"ok_{r['id']}",
                                         type="primary", use_container_width=True):
                                note_f = f"[Admin: {admin_note}]" if admin_note else r.get("note")
                                update_rows("borrow_transactions", {
                                    "condition_in": admin_cond,
                                    "note": note_f,
                                    "status": "คืนแล้ว"
                                }, "id", r["id"])
                                cur_eq = query_table("supplies", select="available_qty,total_qty",
                                                     filters=[("id","eq",r["supply_id"])])
                                if not cur_eq.empty:
                                    cur_a = int(cur_eq.iloc[0]["available_qty"])
                                    cur_t = int(cur_eq.iloc[0]["total_qty"])
                                    qty_ret = int(r["qty"])
                                    if admin_cond == "สูญหาย":
                                        delta, new_stat = 0, "สูญหาย"
                                    elif admin_cond == "ชำรุด":
                                        delta, new_stat = qty_ret, "ชำรุด"
                                    else:
                                        new_avail_r = cur_a + qty_ret
                                        new_stat = "พร้อมใช้" if new_avail_r > 0 else "หมด"
                                        delta = qty_ret
                                    update_rows("supplies", {
                                        "available_qty": cur_a + delta,
                                        "status": new_stat
                                    }, "id", r["supply_id"])
                                load_sidebar_stats.clear()
                                st.success(f"✅ รับคืนแล้ว สภาพ: {admin_cond}")
                                st.rerun()
                        with cc2:
                            if st.button("↩️ ส่งกลับ", key=f"rej_{r['id']}", use_container_width=True):
                                update_rows("borrow_transactions", {
                                    "status": "ยืมอยู่", "return_date": None, "condition_in": None
                                }, "id", r["id"])
                                load_sidebar_stats.clear()
                                st.warning("↩️ ส่งกลับเป็น 'ยืมอยู่' แล้ว")
                                st.rerun()
                    else:
                        st.info("🔒 Login Admin เพื่อยืนยันรับคืน")

    # ── TAB 3: ยกเลิกรายการ (Admin เท่านั้น) ──────────────────────────────────
    with tab3:
        if not is_admin():
            st.warning("🔒 ฟีเจอร์นี้สำหรับ Admin เท่านั้น กรุณา Login ที่ Sidebar")
        else:
            st.markdown(
                '<div class="info-box">🚫 <b>ยกเลิกรายการ</b> — ข้อมูลไม่ถูกลบ '
                'แต่เปลี่ยนสถานะเป็น "ยกเลิก" พร้อมบันทึกเหตุผล ตรวจสอบย้อนหลังได้เสมอ</div>',
                unsafe_allow_html=True)

            # ดึงรายการที่ยกเลิกได้
            df_active_all = load_active_borrows()
            df_pending_all = load_pending_borrows()
            df_cancelable = pd.concat([df_active_all, df_pending_all], ignore_index=True)                             if not df_active_all.empty or not df_pending_all.empty else pd.DataFrame()

            if df_cancelable.empty:
                st.info("✅ ไม่มีรายการที่สามารถยกเลิกได้")
            else:
                st.caption(f"พบ {len(df_cancelable)} รายการที่ยกเลิกได้")
                for _, r in df_cancelable.iterrows():
                    od = overdue_days(r["due_date"])
                    status_color = "#D62828" if r["status"] == "รอตรวจสอบ" else "#52796F"
                    lbl = f"TX#{r['id']} | {r['sup_code']} {r['sup_name']} | {r['br_name']} | [{r['status']}]"

                    with st.expander(lbl):
                        c1, c2 = st.columns([1, 2])
                        with c1:
                            show_image(r.get("sup_img"), size="preview")
                        with c2:
                            st.markdown(
                                f"📦 **{r['sup_code']}** — {r['sup_name']} ({r['qty']} ชิ้น)<br>"
                                f"👤 {r['br_name']} ({r['br_type']})"
                                + (f" | 📞 {r['br_phone']}" if pd.notna(r.get('br_phone')) and r.get('br_phone') else "") + "<br>"
                                f"📅 เบิก {r['borrow_date']} | กำหนดคืน <b>{r['due_date']}</b><br>"
                                f"สถานะ: <b style='color:{status_color};'>{r['status']}</b>"
                                + (f"<br><b style='color:#D62828;'>⚠️ เกิน {od} วัน</b>" if od > 0 else ""),
                                unsafe_allow_html=True)

                        st.markdown("---")
                        cancel_reason = st.text_input(
                            "📝 เหตุผลที่ยกเลิก *",
                            placeholder="เช่น บันทึกผิด, ทดสอบระบบ, ผู้ยืมยกเลิก",
                            key=f"cancel_reason_{r['id']}")

                        if st.button("🚫 ยืนยันยกเลิกรายการนี้", key=f"cancel_{r['id']}",
                                     type="primary", use_container_width=True):
                            if not cancel_reason.strip():
                                st.error("❌ กรุณากรอกเหตุผลที่ยกเลิก")
                            else:
                                try:
                                    # บันทึกสถานะยกเลิก + เหตุผล + วันที่
                                    note_cancel = f"[ยกเลิกโดย Admin | {date.today()} | เหตุผล: {cancel_reason.strip()}]"
                                    update_rows("borrow_transactions", {
                                        "status": "ยกเลิก",
                                        "note": note_cancel
                                    }, "id", r["id"])
                                    # คืน available_qty กลับให้อุปกรณ์
                                    cur_eq = query_table("supplies", select="available_qty,total_qty",
                                                         filters=[("id","eq",r["supply_id"])])
                                    if not cur_eq.empty:
                                        cur_a = int(cur_eq.iloc[0]["available_qty"])
                                        cur_t = int(cur_eq.iloc[0]["total_qty"])
                                        new_a = min(cur_a + int(r["qty"]), cur_t)
                                        new_stat = "พร้อมใช้" if new_a > 0 else "หมด"
                                        update_rows("supplies", {
                                            "available_qty": new_a,
                                            "status": new_stat
                                        }, "id", r["supply_id"])
                                    load_sidebar_stats.clear()
                                    st.success(f"🚫 ยกเลิก TX#{r['id']} เรียบร้อยแล้ว | เหตุผล: {cancel_reason.strip()}")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ เกิดข้อผิดพลาด: {e}")

            # แสดงประวัติที่ยกเลิกแล้ว
            st.markdown('<div class="section-header">📋 ประวัติรายการที่ยกเลิกแล้ว</div>', unsafe_allow_html=True)
            df_cancelled = query_table("borrow_transactions",
                                       select="id,supply_id,borrower_id,qty,borrow_date,due_date,note,status",
                                       filters=[("status","eq","ยกเลิก")],
                                       order=[("id",{"desc":True})])
            if df_cancelled.empty:
                st.info("ยังไม่มีรายการที่ยกเลิก")
            else:
                df_cancelled = df_cancelled.copy()
                df_cancelled["supply_id"]   = df_cancelled["supply_id"].astype(int)
                df_cancelled["borrower_id"] = df_cancelled["borrower_id"].astype(int)
                sup_ids_c = df_cancelled["supply_id"].unique().tolist()
                brr_ids_c = df_cancelled["borrower_id"].unique().tolist()
                df_sc = query_table("supplies", select="id,code,name",
                                    filters=[("id","in_",sup_ids_c)])
                df_bc = query_table("office_borrowers", select="id,name",
                                    filters=[("id","in_",brr_ids_c)])
                if not df_sc.empty:
                    df_sc["id"] = df_sc["id"].astype(int)
                if not df_bc.empty:
                    df_bc["id"] = df_bc["id"].astype(int)
                merged_c = df_cancelled.merge(
                    df_sc.rename(columns={"id":"sid","name":"sname","code":"scode"}),
                    left_on="supply_id", right_on="sid", how="left"
                ).merge(
                    df_bc.rename(columns={"id":"bid","name":"bname"}),
                    left_on="borrower_id", right_on="bid", how="left"
                )
                for _, r in merged_c.iterrows():
                    st.markdown(
                        f'<div class="item-card" style="border-left:4px solid #aaa;opacity:0.75;">'
                        f'🚫 <b>TX#{r["id"]}</b> | {r.get("scode","-")} — {r.get("sname","-")} ({r["qty"]} ชิ้น)<br>'
                        f'👤 {r.get("bname","-")} | 📅 เบิก {r["borrow_date"]}<br>'
                        f'<span style="font-size:0.82rem;color:#888;">{r.get("note","")}</span>'
                        f'</div>', unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: REPORT
# ═════════════════════════════════════════════════════════════════════════════
def page_report():
    st.markdown('<div class="page-title">📊 รายงาน</div>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📋 ประวัติเบิก-คืน","🖊️ ประวัติเบิกสิ้นเปลือง","📦 สรุปคลัง"])

    # ── Tab 1: borrow history ─────────────────────────────────────────────
    with tab1:
        col_d1, col_d2 = st.columns(2)
        date_from = col_d1.date_input("ตั้งแต่", value=date(date.today().year, 1, 1), key="rp_from")
        date_to   = col_d2.date_input("ถึง", value=date.today(), key="rp_to")
        status_f  = st.selectbox("สถานะ", ["ทั้งหมด","ยืมอยู่","คืนแล้ว","รอตรวจสอบ","ยกเลิก"], key="rp_stat")

        filters_r = [("borrow_date","gte",str(date_from)),("borrow_date","lte",str(date_to))]
        if status_f != "ทั้งหมด":
            filters_r.append(("status","eq",status_f))

        df_tx = query_table("borrow_transactions", filters=filters_r, order=[("id",{"desc":True})])
        if not df_tx.empty:
            # cast เป็น int ป้องกัน float key ทำให้ merge ไม่ match
            df_tx["supply_id"]   = df_tx["supply_id"].astype(int)
            df_tx["borrower_id"] = df_tx["borrower_id"].astype(int)
            sup_ids = df_tx["supply_id"].unique().tolist()
            brr_ids = df_tx["borrower_id"].unique().tolist()
            df_s = query_table("supplies", select="id,code,name,group_name",
                               filters=[("id","in_",sup_ids)])
            df_b = query_table("office_borrowers", select="id,name,type,student_id,department,phone",
                               filters=[("id","in_",brr_ids)])
            # cast id ใน lookup tables ด้วย
            if not df_s.empty:
                df_s["id"] = df_s["id"].astype(int)
            if not df_b.empty:
                df_b["id"] = df_b["id"].astype(int)
            merged = df_tx.merge(
                df_s.rename(columns={"id":"sid","name":"sname","code":"scode","group_name":"sgroup"}),
                left_on="supply_id", right_on="sid", how="left"
            ).merge(
                df_b.rename(columns={"id":"bid","name":"bname","type":"btype",
                                     "student_id":"bsid","department":"bdept","phone":"bphone"}),
                left_on="borrower_id", right_on="bid", how="left"
            )
            df_show = pd.DataFrame({
                "TX#": merged["id"], "กลุ่ม": merged["sgroup"],
                "รหัส": merged["scode"], "ชื่ออุปกรณ์": merged["sname"],
                "ผู้ยืม": merged["bname"], "ประเภท": merged["btype"],
                "รหัส/ID": merged["bsid"], "ภาควิชา": merged["bdept"],
                "โทรศัพท์": merged["bphone"], "จำนวน": merged["qty"],
                "วันเบิก": merged["borrow_date"], "กำหนดคืน": merged["due_date"],
                "วันคืน": merged["return_date"] if "return_date" in merged.columns else None,
                "สภาพเบิก": merged["condition_out"] if "condition_out" in merged.columns else None,
                "สภาพคืน": merged["condition_in"] if "condition_in" in merged.columns else None,
                "สถานะ": merged["status"],
                "หมายเหตุ": merged["note"] if "note" in merged.columns else None
            })
        else:
            df_show = pd.DataFrame()

        st.caption(f"พบ {len(df_show)} รายการ")
        st.dataframe(df_show, use_container_width=True, hide_index=True)
        if not df_show.empty and is_admin():
            st.download_button("📥 Export Excel", data=export_excel(df_show, "ประวัติเบิก-คืน"),
                               file_name=f"borrow_{date.today()}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    # ── Tab 2: consumable history ─────────────────────────────────────────
    with tab2:
        col_d3, col_d4 = st.columns(2)
        date_from2 = col_d3.date_input("ตั้งแต่", value=date(date.today().year, 1, 1), key="rp2_from")
        date_to2   = col_d4.date_input("ถึง", value=date.today(), key="rp2_to")
        group_f2   = st.selectbox("กลุ่ม", ["ทั้งหมด"] + list(SUPPLY_GROUPS.keys()), key="rp2_grp")

        filters_c = [("request_date","gte",str(date_from2)),("request_date","lte",str(date_to2))]
        df_con = query_table("consume_transactions", filters=filters_c, order=[("id",{"desc":True})])

        if not df_con.empty:
            sup_ids2 = df_con["supply_id"].unique().tolist()
            df_s2 = query_table("supplies", select="id,code,name,group_name",
                                filters=[("id","in_",sup_ids2)])
            merged2 = df_con.merge(
                df_s2.rename(columns={"id":"sid","name":"sname","code":"scode","group_name":"sgroup"}),
                left_on="supply_id", right_on="sid", how="left"
            )
            if group_f2 != "ทั้งหมด":
                merged2 = merged2[merged2["sgroup"] == group_f2]
            df_show2 = pd.DataFrame({
                "TX#": merged2["id"], "กลุ่ม": merged2["sgroup"],
                "รหัส": merged2["scode"], "ชื่ออุปกรณ์": merged2["sname"],
                "ผู้เบิก": merged2["requester_name"], "ประเภท": merged2["requester_type"],
                "ภาควิชา": merged2["department"],
                "จำนวน": merged2["qty"], "วันเบิก": merged2["request_date"],
                "วัตถุประสงค์": merged2["purpose"] if "purpose" in merged2.columns else None,
                "สถานะ": merged2["status"]
            })
        else:
            df_show2 = pd.DataFrame()

        st.caption(f"พบ {len(df_show2)} รายการ")
        st.dataframe(df_show2, use_container_width=True, hide_index=True)
        if not df_show2.empty and is_admin():
            st.download_button("📥 Export Excel", data=export_excel(df_show2, "ประวัติเบิกสิ้นเปลือง"),
                               file_name=f"consume_{date.today()}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    # ── Tab 3: คลังสรุป ───────────────────────────────────────────────────
    with tab3:
        df_inv = query_table("supplies",
                             select="id,code,name,group_name,unit,total_qty,available_qty,status,min_qty",
                             order=[("group_name",{"desc":False}),("code",{"desc":False})])
        if not df_inv.empty:
            df_inv["กำลังยืม"] = df_inv["total_qty"].astype(int) - df_inv["available_qty"].astype(int)
            df_inv["แจ้งเตือน"] = df_inv.apply(
                lambda row: "⚠️ ต่ำ" if pd.notna(row.get("min_qty")) and int(row["available_qty"]) <= int(row["min_qty"]) else "",
                axis=1)
            df_show3 = df_inv.rename(columns={
                "code":"รหัส","name":"ชื่ออุปกรณ์","group_name":"กลุ่ม",
                "unit":"หน่วย","total_qty":"ทั้งหมด","available_qty":"คงเหลือ","status":"สถานะ"
            })[["รหัส","ชื่ออุปกรณ์","กลุ่ม","หน่วย","ทั้งหมด","คงเหลือ","กำลังยืม","สถานะ","แจ้งเตือน"]]
            st.dataframe(df_show3, use_container_width=True, hide_index=True)
            if is_admin():
                st.download_button("📥 Export Excel สรุปคลัง",
                                   data=export_excel(df_show3, "สรุปคลัง"),
                                   file_name=f"inventory_{date.today()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
        else:
            st.info("ยังไม่มีข้อมูล")

# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def export_excel(df, sheet_name):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill  = PatternFill("solid", fgColor="1b4332")
    hfont  = Font(color="FFFFFF", bold=True, size=11)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill, cell.font, cell.border = hfill, hfont, border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EBF3FB" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border, cell.fill = border, fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 30)
    ws.row_dimensions[1].height = 25
    wb.save(output)
    return output.getvalue()

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: SETTINGS
# ═════════════════════════════════════════════════════════════════════════════
def page_settings():
    st.markdown('<div class="page-title">⚙️ ตั้งค่าระบบ</div>', unsafe_allow_html=True)

    if not is_admin():
        st.warning("🔒 หน้านี้สำหรับ Admin เท่านั้น กรุณา Login ที่ Sidebar")
        return

    tab1, tab2, tab3 = st.tabs(["💾 Backup JSON","📂 Import","🗑️ ล้างข้อมูล"])

    with tab1:
        st.markdown('<div class="section-header">💾 Export Backup</div>', unsafe_allow_html=True)
        if st.button("📦 สร้าง Backup JSON", type="primary", use_container_width=True):
            sup  = query_table("supplies").to_dict(orient="records")
            btx  = query_table("borrow_transactions").to_dict(orient="records")
            ctx  = query_table("consume_transactions").to_dict(orient="records")
            borr = query_table("office_borrowers").to_dict(orient="records")
            backup = {
                "backup_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "version": "office-v1.0",
                "supplies": sup,
                "borrow_transactions": btx,
                "consume_transactions": ctx,
                "office_borrowers": borr
            }
            json_str = json.dumps(backup, ensure_ascii=False, indent=2, default=str)
            fname = f"office_backup_{date.today()}.json"
            st.download_button(f"⬇️ ดาวน์โหลด {fname}", data=json_str.encode("utf-8"),
                               file_name=fname, mime="application/json", use_container_width=True)
            st.success(f"✅ Backup สำเร็จ — อุปกรณ์ {len(sup)} | เบิก-คืน {len(btx)} | สิ้นเปลือง {len(ctx)}")

    with tab2:
        st.markdown('<div class="section-header">📂 Import จาก JSON</div>', unsafe_allow_html=True)
        st.warning("⚠️ Import จะ **เพิ่ม** ข้อมูลเข้าระบบ ไม่ลบของเดิม")
        uploaded = st.file_uploader("เลือกไฟล์ JSON", type=["json"], key="imp_json")
        if uploaded:
            try:
                data = json.loads(uploaded.read().decode("utf-8"))
                st.info(f"📋 Backup วันที่: {data.get('backup_date','?')} | "
                        f"อุปกรณ์: {len(data.get('supplies',[]))} รายการ")

                if st.button("📂 ยืนยัน Import", type="primary", use_container_width=True):
                    imported = 0
                    for s in data.get("supplies", []):
                        try:
                            exist = query_table("supplies", select="id", filters=[("code","eq",s["code"])])
                            if exist.empty:
                                insert_row("supplies", {
                                    "code": s["code"], "name": s["name"],
                                    "group_name": s.get("group_name"),
                                    "group_type": s.get("group_type","consumable"),
                                    "unit": s.get("unit","ชิ้น"),
                                    "total_qty": s.get("total_qty", 1),
                                    "available_qty": s.get("available_qty", 1),
                                    "status": s.get("status","พร้อมใช้"),
                                    "image_url": s.get("image_url"),
                                    "description": s.get("description"),
                                    "min_qty": s.get("min_qty")
                                })
                                imported += 1
                        except Exception:
                            pass
                    load_sidebar_stats.clear()
                    st.success(f"✅ Import {imported} รายการสำเร็จ!")
                    st.rerun()
            except Exception as e:
                st.error(f"❌ ไฟล์ไม่ถูกต้อง: {e}")

    with tab3:
        st.markdown('<div class="section-header">🗑️ ล้างข้อมูล</div>', unsafe_allow_html=True)
        st.error("⚠️ ไม่สามารถกู้คืนได้ แนะนำให้ Backup ก่อนทุกครั้ง!")

        clear_mode = st.selectbox("เลือกประเภทการล้าง", [
            "เลือก...",
            "🔄 รีเซ็ตจำนวน (available = total)",
            "📋 ล้างประวัติการเบิกทั้งหมด",
            "💥 ล้างทุกอย่าง",
        ])

        if clear_mode != "เลือก...":
            # ตรวจสอบรายการ "ยืมอยู่" ก่อนแสดงฟอร์ม
            df_active_chk = query_table("borrow_transactions", select="id",
                                        filters=[("status","eq","ยืมอยู่")])
            n_active_chk = len(df_active_chk)

            if "ล้างประวัติ" in clear_mode or "ล้างทุกอย่าง" in clear_mode:
                if n_active_chk > 0:
                    st.error(f"❌ ไม่สามารถล้างได้ ยังมีอุปกรณ์ที่ยืมอยู่ {n_active_chk} รายการ")
                    st.warning("⚠️ กรุณารับคืนหรือยกเลิกรายการที่ยืมอยู่ทั้งหมดก่อน แล้วค่อยล้างข้อมูล")
                    st.stop()

            st.markdown("**พิมพ์ CONFIRM เพื่อยืนยัน:**")
            confirm = st.text_input("", placeholder="CONFIRM", key="confirm_clear")
            if st.button("🗑️ ดำเนินการ", type="primary", use_container_width=True):
                if confirm != "CONFIRM":
                    st.error("❌ พิมพ์ CONFIRM ให้ถูกต้อง")
                else:
                    try:
                        if "รีเซ็ต" in clear_mode:
                            df_s = query_table("supplies", select="id,total_qty,available_qty,status")
                            for _, r in df_s.iterrows():
                                if int(r["available_qty"]) != int(r["total_qty"]) or r["status"] != "พร้อมใช้":
                                    update_rows("supplies", {
                                        "available_qty": int(r["total_qty"]), "status": "พร้อมใช้"
                                    }, "id", int(r["id"]))
                            st.success("✅ รีเซ็ตเรียบร้อย")
                        elif "ล้างประวัติ" in clear_mode:
                            # reset available_qty ก่อนล้าง
                            df_s = query_table("supplies", select="id,total_qty")
                            for _, r in df_s.iterrows():
                                update_rows("supplies", {
                                    "available_qty": int(r["total_qty"]), "status": "พร้อมใช้"
                                }, "id", int(r["id"]))
                            delete_rows("borrow_transactions", delete_all=True)
                            delete_rows("consume_transactions", delete_all=True)
                            delete_rows("office_borrowers", delete_all=True)
                            st.success("✅ ล้างประวัติเรียบร้อย และรีเซ็ตจำนวนอุปกรณ์แล้ว")
                        elif "ล้างทุกอย่าง" in clear_mode:
                            delete_rows("borrow_transactions", delete_all=True)
                            delete_rows("consume_transactions", delete_all=True)
                            delete_rows("office_borrowers", delete_all=True)
                            delete_rows("supplies", delete_all=True)
                            st.success("✅ ล้างทุกอย่างเรียบร้อย")
                        load_sidebar_stats.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ {e}")

# ─── FOOTER ───────────────────────────────────────────────────────────────────
def footer():
    st.markdown("---")
    st.markdown("""
    <div style="text-align:center;color:#999;font-size:0.8rem;padding:8px 0 16px 0;line-height:1.9;">
        🏢 ระบบบริหารจัดการเบิกอุปกรณ์สำนักงาน<br>
        <b style="color:#2D6A4F;">☁️ Cloud Edition v1.0</b> — Supabase + Cloudinary<br>
        พัฒนาโดย <b style="color:#1b4332;">รศ.ดร.อิทธิพล มีผล</b><br>
        ภาควิชาครุศาสตร์โยธา &nbsp;|&nbsp; คณะครุศาสตร์อุตสาหกรรม<br>
        มหาวิทยาลัยเทคโนโลยีพระจอมเกล้าพระนครเหนือ (KMUTNB)
    </div>
    """, unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    nav()
    page = st.session_state.get("page", "Dashboard")
    if   page == "Dashboard":  page_dashboard()
    elif page == "คลังอุปกรณ์": page_inventory()
    elif page == "เบิกอุปกรณ์": page_request()
    elif page == "คืนอุปกรณ์": page_return()
    elif page == "รายงาน":     page_report()
    elif page == "ตั้งค่า":    page_settings()
    footer()

if __name__ == "__main__":
    main()
