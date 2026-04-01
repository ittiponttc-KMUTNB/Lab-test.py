import streamlit as st
import pandas as pd
import json
import time
from datetime import datetime, date
from zoneinfo import ZoneInfo
_TZ_BKK = ZoneInfo("Asia/Bangkok")
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from supabase import create_client
import cloudinary
import cloudinary.uploader

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ระบบอุปกรณ์ Lab TTC",
    page_icon="🔬",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ─── Secrets ──────────────────────────────────────────────────────────────────
# Streamlit Cloud → Settings → Secrets:
#
# [supabase]
# url = "https://xxxxx.supabase.co"
# key = "eyJhbG..."
#
# [cloudinary]
# cloud_name = "your_cloud_name"
# api_key    = "123456789"
# api_secret = "abcdef..."
#
# [app]
# admin_password = "admin1234"
# logo_url       = "https://res.cloudinary.com/xxx/image/upload/lab_equipment/ttc_logo.png"
# ──────────────────────────────────────────────────────────────────────────────

SUPABASE_URL   = st.secrets["supabase"]["url"]
SUPABASE_KEY   = st.secrets["supabase"]["key"]
ADMIN_PASSWORD = st.secrets["app"]["admin_password"]
LOGO_URL       = st.secrets.get("app", {}).get("logo_url", "")

cloudinary.config(
    cloud_name=st.secrets["cloudinary"]["cloud_name"],
    api_key=st.secrets["cloudinary"]["api_key"],
    api_secret=st.secrets["cloudinary"]["api_secret"],
    secure=True
)

MAX_UPLOAD_MB = 5  # จำกัดขนาดรูปสูงสุด

# ─── Supabase Client ──────────────────────────────────────────────────────────
@st.cache_resource
def get_supabase():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

sb = get_supabase()

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Sarabun', sans-serif;
    font-size: 16px;
    background-color: #ffffff;
}
/* padding ล่างเพื่อไม่ให้ dropdown โดน footer บัง */
section[data-testid="stMain"] > div:first-child {
    padding-bottom: 120px !important;
}
[data-testid="collapsedControl"] { display: none; }
div.stButton > button {
    height: 3rem; font-size: 1rem; border-radius: 10px;
    font-family: 'Sarabun', sans-serif; font-weight: 600;
}
input, textarea { font-size: 16px !important; }
div[data-baseweb="select"] { font-size: 16px; }

div[data-baseweb="select"] > div,
div[data-baseweb="input"] > div {
    background-color: white !important;
    border: 1.5px solid #e8c4bc !important;
    border-radius: 8px !important;
}
div[data-baseweb="select"] > div:hover,
div[data-baseweb="input"] > div:hover { border-color: #7d3020 !important; }

div[data-testid="metric-container"] {
    background: white; border-radius: 12px;
    padding: 14px; border: 1px solid #e8c4bc;
    box-shadow: 0 2px 8px rgba(92,32,24,0.07);
}

/* App Header */
.app-header {
    background: linear-gradient(135deg, #5c2018 0%, #7d3020 100%);
    border-radius: 14px; padding: 16px 18px 12px 18px;
    margin-bottom: 12px; color: white;
    display: flex; align-items: center; gap: 12px;
}
.app-header h2 { margin: 0; font-size: 1.1rem; font-weight: 700; color: white; }
.app-header p  { margin: 0; font-size: 0.78rem; color: #f0c4bc; }

/* Role badge */
.role-badge { display: inline-flex; align-items: center; gap: 5px;
    padding: 4px 12px; border-radius: 20px; font-size: 0.8rem; font-weight: 700;
    margin-bottom: 12px; }
.role-user  { background: #fdf2f0; color: #5c2018; border: 1px solid #e8c4bc; }
.role-admin { background: #fff3cd; color: #856404; border: 1px solid #ffc107; }

/* Section header */
.section-header {
    font-size: 1rem; font-weight: 700; color: #5c2018;
    margin: 16px 0 8px 0; padding: 6px 10px;
    background: #fdf2f0; border-left: 4px solid #7d3020;
    border-radius: 0 6px 6px 0;
}

/* Page title */
.page-title {
    font-size: 1.3rem; font-weight: 700; color: #5c2018;
    margin: 0 0 14px 0; padding-bottom: 8px;
    border-bottom: 3px solid #7d3020;
}

/* Cards */
.eq-card {
    background: white; border-radius: 10px; padding: 12px 14px;
    margin-bottom: 10px; border: 1px solid #e8c4bc;
    box-shadow: 0 2px 6px rgba(92,32,24,0.06); line-height: 1.7;
}
.overdue-alert {
    background: #fff3cd; border-left: 4px solid #ff6b6b;
    border-radius: 0 8px 8px 0; padding: 10px 14px; margin: 6px 0; line-height: 1.7;
}

/* Info box */
.info-box {
    background: #fdf2f0; border: 1px solid #e8c4bc;
    border-radius: 8px; padding: 10px 14px; margin: 8px 0;
    font-size: 0.88rem; color: #5c2018;
}

/* Badge */
.badge {
    display: inline-block; padding: 2px 10px; border-radius: 20px;
    font-size: 0.78rem; font-weight: 600; color: white;
}

/* Tabs — Pill style */
div[data-testid="stTabs"] > div:first-child {
    background: #fdf2f0; border-radius: 12px; padding: 5px; gap: 6px;
}
div[data-testid="stTabs"] div[role="tablist"] { border-bottom: none !important; gap: 6px !important; }
div[data-testid="stTabs"] button[role="tab"] {
    border-radius: 8px !important; font-size: 0.92rem !important; font-weight: 600 !important;
    padding: 9px 14px !important; color: #5c2018 !important;
    background: white !important; border: 1.5px solid #e8c4bc !important;
    transition: all 0.15s ease !important;
}
div[data-testid="stTabs"] button[role="tab"]:hover {
    background: #fdf2f0 !important; border-color: #7d3020 !important;
}
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    background: #7d3020 !important; color: white !important;
    border-color: #7d3020 !important; box-shadow: 0 2px 8px rgba(125,48,32,0.30) !important;
}

/* Quick action buttons — เบิก=เลือดหมู, คืน=น้ำเงิน */
/* จำกัด scope ด้วย class .home-cards เพื่อไม่กระทบ nav */
div.home-cards div[data-testid="stHorizontalBlock"] div.stButton button {
    height: 130px !important; border-radius: 14px !important;
    font-size: 1rem !important; font-weight: 700 !important;
    white-space: pre-line !important; line-height: 1.6 !important;
    padding: 16px 12px !important; border: none !important;
    color: white !important;
    box-shadow: 0 4px 14px rgba(0,0,0,0.15) !important;
    transition: all 0.18s ease !important;
}
div.home-cards div[data-testid="stHorizontalBlock"] > div:nth-child(1) div.stButton button {
    background: #7d3020 !important;
}
div.home-cards div[data-testid="stHorizontalBlock"] > div:nth-child(1) div.stButton button:hover {
    background: #5c2018 !important;
    transform: translateY(-3px) !important;
    box-shadow: 0 6px 20px rgba(92,32,24,0.35) !important;
}
div.home-cards div[data-testid="stHorizontalBlock"] > div:nth-child(2) div.stButton button {
    background: #1a5276 !important;
}
div.home-cards div[data-testid="stHorizontalBlock"] > div:nth-child(2) div.stButton button:hover {
    background: #0c2461 !important;
    transform: translateY(-3px) !important;
    box-shadow: 0 6px 20px rgba(26,82,118,0.30) !important;
}
div.home-cards div[data-testid="stHorizontalBlock"] div.stButton button:active {
    transform: scale(0.97) !important;
}

/* Step wizard */
.step-bar { display: flex; align-items: center; margin-bottom: 16px; gap: 0; }
.step-item { display: flex; flex-direction: column; align-items: center; flex: 1; }
.step-circle {
    width: 32px; height: 32px; border-radius: 50%; display: flex;
    align-items: center; justify-content: center;
    font-size: 0.9rem; font-weight: 700; border: 2px solid #e8c4bc;
    background: white; color: #aaa;
}
.step-circle.done   { background: #7d3020; border-color: #7d3020; color: white; }
.step-circle.active { background: #5c2018; border-color: #5c2018; color: white; }
.step-label { font-size: 0.7rem; color: #888; margin-top: 3px; }
.step-label.active { color: #5c2018; font-weight: 700; }
.step-line { flex: 1; height: 2px; background: #e8c4bc; }
.step-line.done { background: #7d3020; }

/* Admin sidebar */
.admin-panel-title {
    font-size: 0.85rem; font-weight: 700; color: #856404;
    padding: 6px 0; border-bottom: 1px solid #ffc107; margin-bottom: 8px;
}
</style>
""", unsafe_allow_html=True)

# ─── ADMIN AUTH ─────────────────────────────────────────────────────────────
def is_admin():
    return st.session_state.get("is_admin", False)

def admin_login_widget():
    if is_admin():
        st.sidebar.markdown('<div class="admin-panel-title">🔓 Admin Mode เปิดอยู่</div>', unsafe_allow_html=True)
        if st.sidebar.button("🔒 ออกจาก Admin", use_container_width=True):
            st.session_state.is_admin = False
            st.rerun()
        st.sidebar.markdown("---")
        admin_pages = [
            ("Dashboard","🏠","ภาพรวม"), ("อุปกรณ์","📦","จัดการ"),
            ("คืน","✅","ตรวจสอบ"), ("รายงาน","📋","รายงาน"), ("ตั้งค่า","⚙️","ตั้งค่า"),
        ]
        for pname, icon, label in admin_pages:
            active = st.session_state.get("page") == pname
            if st.sidebar.button(f"{icon} {label}", key=f"sidebar_nav_{pname}",
                                  use_container_width=True,
                                  type="primary" if active else "secondary"):
                st.session_state.page = pname
                st.rerun()
    else:
        with st.sidebar.expander("🔒 Admin Login", expanded=False):
            pwd = st.text_input("รหัสผ่าน Admin", type="password", key="admin_pwd_input")
            if st.button("เข้าสู่ระบบ", use_container_width=True, key="admin_login_btn"):
                if pwd == ADMIN_PASSWORD:
                    st.session_state.is_admin = True
                    st.session_state.page = "Dashboard"
                    st.rerun()
                else:
                    st.error("รหัสผ่านไม่ถูกต้อง")

# ═════════════════════════════════════════════════════════════════════════════
# DATABASE LAYER — Supabase with retry & error handling          [FIX #5, #7]
# ═════════════════════════════════════════════════════════════════════════════
def _sb_retry(func, retries=2, delay=1):
    """Retry wrapper สำหรับ Supabase calls"""
    for attempt in range(retries + 1):
        try:
            return func()
        except Exception as e:
            if attempt == retries:
                raise e
            time.sleep(delay)

def query_table(table, select="*", filters=None, order=None, limit=None):
    """Query Supabase table → DataFrame (with retry)"""
    def _do():
        q = sb.table(table).select(select)
        if filters:
            for col, op, val in filters:
                if op == "in_":
                    q = q.in_(col, val)
                else:
                    q = getattr(q, op if op != "is" else "is_")(col, val)
        if order:
            for col_name, opts in order:
                q = q.order(col_name, **opts)
        if limit:
            q = q.limit(limit)
        resp = q.execute()
        return pd.DataFrame(resp.data) if resp.data else pd.DataFrame()
    return _sb_retry(_do)

def insert_row(table, data):
    """Insert → return inserted row dict"""
    def _do():
        resp = sb.table(table).insert(data).execute()
        return resp.data[0] if resp.data else None
    return _sb_retry(_do)

def update_rows(table, data, match_col, match_val):
    """Update row(s) by single column match"""
    def _do():
        resp = sb.table(table).update(data).eq(match_col, match_val).execute()
        return resp.data
    return _sb_retry(_do)

def delete_rows(table, match_col=None, match_val=None, delete_all=False):
    """Delete row(s) — delete_all ใช้ gt id 0"""
    def _do():
        if delete_all:
            resp = sb.table(table).delete().gt("id", 0).execute()
        else:
            resp = sb.table(table).delete().eq(match_col, match_val).execute()
        return resp.data
    return _sb_retry(_do)


def clear_all_cache():
    """Clear cache ทั้งหมดหลัง write operation"""
    load_sidebar_stats.clear()
    load_active_transactions_enriched.clear()
    load_pending_transactions_enriched.clear()

def batch_reset_equipment():
    """รีเซ็ตทุกอุปกรณ์ — ดึงมาแล้ว update เฉพาะตัวที่ต้องเปลี่ยน [FIX #4]"""
    df = query_table("equipment", select="id,total_qty,available_qty,status")
    if df.empty:
        return
    need_update = df[(df["available_qty"] != df["total_qty"]) | (df["status"] != "พร้อมใช้")]
    for _, r in need_update.iterrows():
        update_rows("equipment",
                     {"available_qty": int(r["total_qty"]), "status": "พร้อมใช้"},
                     "id", int(r["id"]))

# ═════════════════════════════════════════════════════════════════════════════
# DATA LOADING — Batch fetch + merge (eliminates N+1)            [FIX #1, #6]
# ═════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=10, show_spinner=False)
def load_sidebar_stats():
    """Cache sidebar stats 30 วินาที"""
    df_eq = query_table("equipment", select="id,available_qty,is_consumable,min_qty")
    n_eq = len(df_eq)
    avail = int(df_eq["available_qty"].sum()) if not df_eq.empty else 0

    today_str = str(date.today())
    df_active = query_table("transactions", select="id,due_date",
                            filters=[("status", "eq", "ยืมอยู่")])
    n_borr = len(df_active)
    n_over = len(df_active[df_active["due_date"] < today_str]) if not df_active.empty else 0

    # นับวัสดุสิ้นเปลืองใกล้หมด (available_qty <= min_qty)
    n_low = 0
    if not df_eq.empty:
        df_cons = df_eq[df_eq["is_consumable"] == True]
        if not df_cons.empty:
            n_low = len(df_cons[df_cons["available_qty"] <= df_cons["min_qty"].fillna(0)])

    return n_eq, avail, n_borr, n_over, n_low

@st.cache_data(ttl=10, show_spinner=False)
def load_active_transactions_enriched():
    """[FIX #1] ดึง transactions + equipment + borrowers ทีเดียว แล้ว merge"""
    df_tx = query_table("transactions",
                        select="id,equipment_id,borrower_id,qty,borrow_date,due_date,condition_out,note,return_date,condition_in,status",
                        filters=[("status", "eq", "ยืมอยู่")],
                        order=[("due_date", {"desc": False})])
    if df_tx.empty:
        return pd.DataFrame()
    return _enrich_transactions(df_tx)

@st.cache_data(ttl=10, show_spinner=False)
def load_pending_transactions_enriched():
    """[FIX #1] ดึง transactions รอตรวจสอบ + merge"""
    df_tx = query_table("transactions",
                        select="id,equipment_id,borrower_id,qty,borrow_date,due_date,return_date,condition_out,condition_in,note,status",
                        filters=[("status", "eq", "รอตรวจสอบ")],
                        order=[("return_date", {"desc": False})])
    if df_tx.empty:
        return pd.DataFrame()
    return _enrich_transactions(df_tx)

def _enrich_transactions(df_tx):
    """Batch fetch equipment + borrowers เฉพาะ id ที่ต้องการ แล้ว merge"""
    eq_ids = df_tx["equipment_id"].unique().tolist()
    br_ids = df_tx["borrower_id"].unique().tolist()

    # ดึงเฉพาะ id ที่เกี่ยวข้อง (ไม่ดึงทั้ง table)
    df_eq = query_table("equipment", select="id,code,name,image_url,category",
                        filters=[("id", "in_", eq_ids)]) if eq_ids else pd.DataFrame()

    df_br = query_table("borrowers", select="id,name,type,phone,student_id,department",
                        filters=[("id", "in_", br_ids)]) if br_ids else pd.DataFrame()

    # Merge: transactions ← equipment ← borrowers
    merged = df_tx.merge(
        df_eq.rename(columns={"id": "eq_id", "name": "eq_name", "image_url": "eq_image_url",
                               "code": "eq_code", "category": "eq_category"}),
        left_on="equipment_id", right_on="eq_id", how="left"
    ).merge(
        df_br.rename(columns={"id": "br_id", "name": "br_name", "type": "br_type",
                               "phone": "br_phone", "student_id": "br_student_id",
                               "department": "br_department"}),
        left_on="borrower_id", right_on="br_id", how="left"
    )
    return merged

def load_report_data(date_from, date_to, status_filter):
    """[FIX #1] รายงาน — batch fetch + merge"""
    filters = [("borrow_date", "gte", str(date_from)), ("borrow_date", "lte", str(date_to))]
    if status_filter != "ทั้งหมด":
        filters.append(("status", "eq", status_filter))

    df_tx = query_table("transactions", filters=filters, order=[("id", {"desc": True})])
    if df_tx.empty:
        return pd.DataFrame()

    df_eq = query_table("equipment", select="id,code,name",
                        filters=[("id", "in_", df_tx["equipment_id"].unique().tolist())])
    df_br = query_table("borrowers", select="id,name,type,student_id,department,phone",
                        filters=[("id", "in_", df_tx["borrower_id"].unique().tolist())])

    merged = df_tx.merge(
        df_eq.rename(columns={"id": "eq_id", "name": "eq_name", "code": "eq_code"}),
        left_on="equipment_id", right_on="eq_id", how="left"
    ).merge(
        df_br.rename(columns={"id": "br_id", "name": "br_name", "type": "br_type",
                               "student_id": "br_student_id", "department": "br_department",
                               "phone": "br_phone"}),
        left_on="borrower_id", right_on="br_id", how="left"
    )

    report = pd.DataFrame({
        "TX#": merged["id"],
        "รหัส": merged["eq_code"],
        "อุปกรณ์": merged["eq_name"],
        "ผู้เบิก": merged["br_name"],
        "ประเภท": merged["br_type"],
        "รหัส/ID": merged["br_student_id"],
        "ภาควิชา": merged["br_department"],
        "โทรศัพท์": merged["br_phone"],
        "จำนวน": merged["qty"],
        "วันเบิก": merged["borrow_date"],
        "กำหนดคืน": merged["due_date"],
        "วันคืน": merged["return_date"],
        "สภาพตอนเบิก": merged["condition_out"],
        "สภาพตอนคืน": merged["condition_in"],
        "สถานะ": merged["status"],
        "หมายเหตุ": merged["note"]
    })
    return report

def load_equipment_summary():
    """สรุปอุปกรณ์ — นับ tx_count ด้วย batch"""
    df_eq = query_table("equipment",
                        select="id,code,name,category,total_qty,available_qty,status,is_consumable,min_qty",
                        order=[("code", {"desc": False})])
    if df_eq.empty:
        return pd.DataFrame()

    # นับจำนวนครั้งที่เบิกทั้งหมดใน 1 query
    df_tx = query_table("transactions", select="id,equipment_id")
    if not df_tx.empty:
        tx_counts = df_tx.groupby("equipment_id").size().reset_index(name="ครั้งที่เบิก")
    else:
        tx_counts = pd.DataFrame(columns=["equipment_id", "ครั้งที่เบิก"])

    df_eq = df_eq.merge(tx_counts, left_on="id", right_on="equipment_id", how="left")
    df_eq["ครั้งที่เบิก"] = df_eq["ครั้งที่เบิก"].fillna(0).astype(int)
    df_eq["กำลังยืม"] = df_eq["total_qty"].astype(int) - df_eq["available_qty"].astype(int)
    df_eq["ประเภท"] = df_eq["is_consumable"].apply(lambda x: "วัสดุสิ้นเปลือง" if x else "อุปกรณ์")

    return df_eq[["code", "name", "category", "ประเภท", "total_qty", "available_qty",
                   "กำลังยืม", "status", "ครั้งที่เบิก"]].rename(columns={
        "code": "รหัส", "name": "ชื่ออุปกรณ์", "category": "หมวดหมู่",
        "total_qty": "ทั้งหมด", "available_qty": "พร้อมใช้", "status": "สถานะ"
    })

# ═════════════════════════════════════════════════════════════════════════════
# CLOUDINARY — Upload with validation + Optimized URL            [FIX #2, #8]
# ═════════════════════════════════════════════════════════════════════════════
def upload_image(file_obj, public_id):
    """[FIX #8] Upload with size validation + auto optimization"""
    # ตรวจขนาดไฟล์
    file_size_mb = len(file_obj.getvalue()) / (1024 * 1024)
    if file_size_mb > MAX_UPLOAD_MB:
        st.error(f"❌ ไฟล์ใหญ่เกิน {MAX_UPLOAD_MB} MB (ไฟล์นี้ {file_size_mb:.1f} MB) กรุณาลดขนาดก่อนอัพโหลด")
        return None
    try:
        result = cloudinary.uploader.upload(
            file_obj,
            public_id=public_id,
            overwrite=True,
            folder="lab_equipment",
            resource_type="image",
            transformation=[
                {"width": 1200, "crop": "limit"},          # จำกัดสูงสุด 1200px
                {"quality": "auto", "fetch_format": "auto"} # optimize อัตโนมัติ
            ]
        )
        return result.get("secure_url")
    except Exception as e:
        st.error(f"❌ อัพโหลดรูปไม่สำเร็จ: {e}")
        return None

def optimized_url(original_url, width=400, height=300, crop="fill"):
    """[FIX #2] สร้าง URL ที่ optimize แล้วจาก Cloudinary URL ต้นฉบับ"""
    if not original_url or not isinstance(original_url, str) or "cloudinary" not in original_url:
        return original_url
    return original_url.replace(
        "/upload/", f"/upload/w_{width},h_{height},c_{crop},q_auto,f_auto/"
    )

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def show_image(image_url, width=100, size="preview"):
    """แสดงรูปพร้อม Cloudinary optimization — ใช้ c_pad ไม่ตัดรูป"""
    SIZES = {"thumb": (200, 150), "preview": (400, 300), "full": (800, 600)}
    w_img, h_img = SIZES.get(size, (400, 300))
    opt_url = optimized_url(image_url, w_img, h_img, crop="pad")

    w_style = f"{width}px" if isinstance(width, int) else width
    if opt_url and isinstance(opt_url, str) and opt_url.startswith("http"):
        st.markdown(
            f'<img src="{opt_url}" '
            f'style="width:{w_style};max-width:100%;border-radius:8px;'
            f'border:1px solid #ddd;object-fit:contain;" loading="lazy">',
            unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div style="width:{w_style};height:120px;background:#f0f2f6;'
            f'border-radius:8px;display:flex;align-items:center;justify-content:center;'
            f'font-size:2rem;border:1px solid #ddd;">📦</div>',
            unsafe_allow_html=True)

STATUS_COLOR = {
    "พร้อมใช้": "#28a745", "ยืมออก": "#ffc107",
    "ชำรุด":    "#dc3545", "สูญหาย": "#6c757d"
}

def badge(text, color):
    return f'<span class="badge" style="background:{color};">{text}</span>'

def overdue_days(due_str):
    try:
        d = datetime.strptime(str(due_str), "%Y-%m-%d").date()
        delta = (date.today() - d).days
        return max(delta, 0)
    except (ValueError, TypeError):
        return 0

# ═════════════════════════════════════════════════════════════════════════════
# NAVIGATION + HEADER                                           [FIX #3, #9]
# ═════════════════════════════════════════════════════════════════════════════
# ═════════════════════════════════════════════════════════════════════════════
# NAVIGATION + HEADER
# ═════════════════════════════════════════════════════════════════════════════
def render_header():
    logo_html = ""
    if LOGO_URL:
        logo_opt = optimized_url(LOGO_URL, 100, 100, "fit") if "cloudinary" in LOGO_URL else LOGO_URL
        logo_html = f'<img src="{logo_opt}" style="width:72px;height:auto;border-radius:10px;" loading="lazy">'
    else:
        logo_html = '<span style="font-size:3rem;">🔬</span>'

    role_badge = (
        '<span class="role-badge role-admin">🔑 Admin Mode</span>'
        if is_admin() else
        '<span class="role-badge role-user">👤 ผู้ใช้ทั่วไป</span>'
    )
    st.markdown(
        f'<div class="app-header">'
        f'<div style="min-width:80px;">{logo_html}</div>'
        f'<div><h2>ระบบอุปกรณ์ Lab TTC</h2>'
        f'<p>ภาควิชาครุศาสตร์โยธา — มจพ.</p></div></div>',
        unsafe_allow_html=True
    )
    st.markdown(role_badge, unsafe_allow_html=True)


def render_top_nav():
    page = st.session_state.get("page", "หน้าหลัก")

    if is_admin():
        try:
            df_pend = query_table("transactions", select="id", filters=[("status","eq","รอตรวจสอบ")])
            n_p = len(df_pend)
        except Exception:
            n_p = 0
        badge_p = f" ({n_p})" if n_p > 0 else ""

        nav_items = [
            ("หน้าหลัก", "🏠", "หน้าหลัก"),
            ("เบิก",      "➕", "เบิก"),
            ("คืน",       "✅", f"ตรวจสอบ{badge_p}"),
            ("Dashboard", "📊", "ภาพรวม"),
            ("อุปกรณ์",   "📦", "คลัง"),
        ]
    else:
        if page != "หน้าหลัก":
            if st.button("🏠 หน้าหลัก", key="tnav_home_user", use_container_width=False):
                st.session_state.page = "หน้าหลัก"
                st.rerun()
        return

    cols = st.columns(len(nav_items))
    for i, (pname, icon, label) in enumerate(nav_items):
        with cols[i]:
            if st.button(f"{icon}\n{label}", key=f"tnav_{pname}",
                         use_container_width=True,
                         type="primary" if page == pname else "secondary"):
                st.session_state.page = pname
                st.rerun()
    st.markdown('<div style="margin-bottom:8px;"></div>', unsafe_allow_html=True)


def nav():
    if "page" not in st.session_state:
        st.session_state.page = "หน้าหลัก"

    with st.sidebar:
        st.markdown("## 🔬 ระบบอุปกรณ์ Lab")
        admin_login_widget()
        if is_admin():
            st.markdown("---")
            try:
                n_eq, avail, n_borr, n_over, n_low = load_sidebar_stats()
                st.metric("📦 อุปกรณ์", n_eq)
                c1, c2 = st.columns(2)
                c1.metric("✅ พร้อมใช้", avail)
                c2.metric("🔄 ยืมอยู่", n_borr)
                if n_over > 0:
                    st.error(f"⚠️ เกินกำหนด {n_over} รายการ")
                if n_low > 0:
                    st.warning(f"📦 ใกล้หมด {n_low} รายการ")
            except Exception:
                st.caption("⏳ กำลังโหลด...")

    render_header()
    render_top_nav()


# ═════════════════════════════════════════════════════════════════════════════
# PAGE: หน้าหลัก
# ═════════════════════════════════════════════════════════════════════════════
def page_home():
    try:
        n_eq, avail, n_borr, n_over, n_low = load_sidebar_stats()
        c1, c2, c3 = st.columns(3)
        c1.metric("📦 อุปกรณ์", n_eq)
        c2.metric("✅ พร้อมใช้", avail)
        c3.metric("🔄 ยืมอยู่", n_borr)
        if n_over > 0:
            st.error(f"⚠️ มีอุปกรณ์เกินกำหนดคืน {n_over} รายการ!")
        if n_low > 0:
            st.warning(f"📦 วัสดุสิ้นเปลืองใกล้หมด {n_low} รายการ!")
    except Exception:
        st.warning("⏳ กำลังโหลดข้อมูล...")

    st.markdown("---")
    st.markdown('<div class="section-header">🚀 ทำรายการ</div>', unsafe_allow_html=True)
    st.markdown('<div class="home-cards">', unsafe_allow_html=True)
    col_a, col_b = st.columns(2)
    with col_a:
        if st.button("➕\nเบิกอุปกรณ์\nอุปกรณ์ / วัสดุสิ้นเปลือง",
                     key="home_to_borrow", use_container_width=True):
            st.session_state.page = "เบิก"
            st.rerun()
    with col_b:
        if st.button("✅\nคืนอุปกรณ์\nแจ้งนำอุปกรณ์มาคืน",
                     key="home_to_return", use_container_width=True):
            st.session_state.page = "คืน"
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    if not is_admin():
        st.markdown("""
        <div class="info-box" style="margin-top:16px;">
        📌 <b>คำแนะนำการใช้งาน</b><br>
        • <b>เบิกอุปกรณ์</b> — เลือกอุปกรณ์ กรอกชื่อ กดยืนยัน<br>
        • <b>คืนอุปกรณ์</b> — ค้นหารายการที่ยืม กดแจ้งคืน รอ Admin ตรวจสอบ<br>
        • สอบถามเพิ่มเติมติดต่อ <b>เจ้าหน้าที่ห้อง Lab</b>
        </div>
        """, unsafe_allow_html=True)
        st.markdown(
            '<div style="text-align:center;color:#888;font-size:0.8rem;margin-top:12px;">' +
            '🔒 Admin Login ที่เมนู ☰ มุมซ้ายบน</div>',
            unsafe_allow_html=True)
    else:
        st.markdown("---")
        st.markdown('<div class="section-header">⚡ Admin Shortcuts</div>', unsafe_allow_html=True)
        cx, cy, cz = st.columns(3)
        with cx:
            if st.button("📊 ภาพรวม", use_container_width=True, key="home_dash"):
                st.session_state.page = "Dashboard"
                st.rerun()
        with cy:
            if st.button("✅ ตรวจสอบการคืน", use_container_width=True, key="home_ret"):
                st.session_state.page = "คืน"
                st.rerun()
        with cz:
            if st.button("📦 จัดการอุปกรณ์", use_container_width=True, key="home_eq"):
                st.session_state.page = "อุปกรณ์"
                st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: DASHBOARD                                               [FIX #1, #6]
# ═════════════════════════════════════════════════════════════════════════════
def page_dashboard():
    st.markdown('<div class="page-title">🏠 ภาพรวมระบบ Lab</div>', unsafe_allow_html=True)

    # Stats — ใช้ cached sidebar stats
    try:
        n_eq, available, active_tx, overdue_count, n_low = load_sidebar_stats()
    except Exception:
        st.error("❌ ไม่สามารถเชื่อมต่อ Supabase ได้ กรุณาตรวจสอบการตั้งค่า")
        return

    c1, c2 = st.columns(2)
    c1.metric("📦 อุปกรณ์ทั้งหมด", n_eq)
    c2.metric("✅ พร้อมใช้", available)
    c1.metric("🔄 กำลังยืม", active_tx)
    c2.metric("⚠️ เกินกำหนด", overdue_count)

    # [FIX #1] ดึง enriched transactions ทีเดียว
    df = load_active_transactions_enriched()

    if overdue_count > 0 and not df.empty:
        st.error(f"⚠️ มีอุปกรณ์เกินกำหนดคืน {overdue_count} รายการ!")
        today_str = str(date.today())
        df_od = df[df["due_date"] < today_str].copy()
        df_od["days_over"] = df_od["due_date"].apply(overdue_days)
        df_od = df_od.sort_values("days_over", ascending=False)
        for _, r in df_od.iterrows():
            phone_str = f"📞 {r['br_phone']}" if pd.notna(r.get('br_phone')) and r.get('br_phone') else ""
            st.markdown(f"""
            <div class="overdue-alert">
                🔴 <b>{r['eq_code']} — {r['eq_name']}</b><br>
                👤 {r['br_name']} {phone_str}<br>
                📅 กำหนดคืน {r['due_date']} &nbsp;
                <b style="color:red;">เกิน {r['days_over']} วัน</b>
            </div>""", unsafe_allow_html=True)

    # ── แจ้งเตือนวัสดุสิ้นเปลืองใกล้หมด ──────────────────────────────────
    if n_low > 0:
        st.warning(f"📦 วัสดุสิ้นเปลืองใกล้หมด {n_low} รายการ!")
        df_low = query_table("equipment",
                             select="code,name,available_qty,min_qty",
                             filters=[("is_consumable","eq",True)])
        if not df_low.empty:
            df_low = df_low[df_low["available_qty"] <= df_low["min_qty"].fillna(0)]
            for _, r in df_low.iterrows():
                color = "#dc3545" if r["available_qty"] == 0 else "#ffc107"
                st.markdown(
                    f'<div class="eq-card" style="border-left:4px solid {color};">'
                    f'📦 <b>{r["code"]}</b> — {r["name"]}<br>'
                    f'คงเหลือ: <b style="color:{color};">{r["available_qty"]}</b>'
                    f' / แจ้งเตือนเมื่อ ≤ {int(r["min_qty"])}</div>',
                    unsafe_allow_html=True)

    st.markdown('<div class="section-header">📋 รายการที่กำลังยืมอยู่</div>', unsafe_allow_html=True)
    if df.empty:
        st.info("ไม่มีรายการยืมในขณะนี้ ✅")
    else:
        for _, r in df.iterrows():
            od = overdue_days(r["due_date"])
            bc = "#ff6b6b" if od > 0 else "#28a745"
            img_url = optimized_url(r.get("eq_image_url"), 100, 75)
            st.markdown(f"""
            <div class="eq-card" style="border-left:4px solid {bc};">
                <b>{r['eq_code']}</b> — {r['eq_name']}<br>
                👤 {r['br_name']} ({r['br_type']}) &nbsp; 📦 {r['qty']}<br>
                📅 เบิก {r['borrow_date']} | คืน <b>{r['due_date']}</b>
                {"&nbsp;<b style='color:red;'>⚠️ เกิน "+str(od)+" วัน</b>" if od > 0 else ""}
            </div>""", unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: EQUIPMENT                                               [FIX #1, #2]
# ═════════════════════════════════════════════════════════════════════════════
def page_equipment():
    st.markdown('<div class="page-title">📦 รายการอุปกรณ์ Lab</div>', unsafe_allow_html=True)

    search = st.text_input("🔍 ค้นหา ชื่อ / รหัส / หมวดหมู่", placeholder="พิมพ์เพื่อค้นหา...")

    df_all_eq = query_table("equipment",
                            select="id,code,name,category,total_qty,available_qty,status,image_url,description,is_consumable,min_qty",
                            order=[("code", {"desc": False})])
    cats = sorted(df_all_eq["category"].dropna().unique().tolist()) if not df_all_eq.empty else []
    cat_filter = st.selectbox("📂 หมวดหมู่", ["ทั้งหมด"] + cats)

    df = df_all_eq.copy()
    if search:
        mask = (df["name"].str.contains(search, case=False, na=False) |
                df["code"].str.contains(search, case=False, na=False) |
                df["category"].str.contains(search, case=False, na=False))
        df = df[mask]
    if cat_filter != "ทั้งหมด":
        df = df[df["category"] == cat_filter]

    st.caption(f"พบ {len(df)} รายการ")

    # [FIX #1] batch fetch ผู้ยืมล่าสุดทั้งหมด
    df_last_tx = pd.DataFrame()
    df_br_all = pd.DataFrame()
    if not df.empty:
        df_all_tx = query_table("transactions", select="id,equipment_id,borrower_id,borrow_date,created_at",
                                order=[("created_at", {"desc": True})])
        if not df_all_tx.empty:
            df_last_tx = df_all_tx.drop_duplicates(subset=["equipment_id"], keep="first")
            br_ids = df_last_tx["borrower_id"].unique().tolist()
            df_br_all = query_table("borrowers", select="id,name")
            if not df_br_all.empty:
                df_br_all = df_br_all[df_br_all["id"].isin(br_ids)]

    for _, r in df.iterrows():
        with st.expander(f"{r['code']} — {r['name']}"):
            c1, c2 = st.columns([1, 2])
            with c1:
                show_image(r.get("image_url"), width="100%", size="preview")
            with c2:
                st.markdown(
                    f"**หมวด:** {r['category'] or '-'}<br>"
                    f"**ทั้งหมด:** {r['total_qty']} | **พร้อมใช้:** {r['available_qty']}<br>"
                    f"**สถานะ:** {badge(r['status'], STATUS_COLOR.get(r['status'],'#888'))}"
                    + (f" {badge('วัสดุสิ้นเปลือง', '#17a2b8')}" if r.get('is_consumable') else ""),
                    unsafe_allow_html=True)
                if r.get("description"):
                    st.caption(r["description"])

            # ผู้ยืมล่าสุด — ใช้ข้อมูลที่ fetch มาแล้ว (ไม่ query ใน loop)
            if not df_last_tx.empty:
                lt = df_last_tx[df_last_tx["equipment_id"] == r["id"]]
                if not lt.empty and not df_br_all.empty:
                    br = df_br_all[df_br_all["id"] == lt.iloc[0]["borrower_id"]]
                    if not br.empty:
                        st.caption(f"👤 ผู้ยืมล่าสุด: **{br.iloc[0]['name']}** ({lt.iloc[0]['borrow_date']})")

            if is_admin():
                btn_col1, btn_col2 = st.columns(2)
                with btn_col1:
                    if st.button("✏️ แก้ไข", key=f"edit_{r['id']}", use_container_width=True):
                        opts_tmp = ["➕ เพิ่มใหม่"] + [f"{x['code']} — {x['name']}" for _, x in df_all_eq.iterrows()]
                        matched_tmp = [o for o in opts_tmp if o.startswith(r["code"] + " —")]
                        if matched_tmp:
                            st.session_state["_next_sel"] = matched_tmp[0]
                        st.rerun()
                with btn_col2:
                    if st.button("🗑️ ลบ", key=f"del_{r['id']}", use_container_width=True):
                        active = query_table("transactions", select="id",
                                             filters=[("equipment_id","eq",r["id"]),("status","eq","ยืมอยู่")])
                        if len(active) > 0:
                            st.error("ไม่สามารถลบได้ มีการยืมอยู่")
                        else:
                            delete_rows("equipment", "id", r["id"])
                            st.success("ลบแล้ว")
                            clear_all_cache()
                            st.rerun()

    # ── Admin: เพิ่ม/แก้ไขอุปกรณ์ ──────────────────────────────────────────
    if is_admin():
        st.markdown('<div class="section-header">➕ เพิ่ม / แก้ไขอุปกรณ์</div>', unsafe_allow_html=True)

        options = ["➕ เพิ่มใหม่"] + [f"{r['code']} — {r['name']}" for _, r in df_all_eq.iterrows()]

        if "_next_sel" in st.session_state:
            st.session_state["eq_edit_sel"] = st.session_state.pop("_next_sel")

        choice = st.selectbox("เลือกรายการ", options, key="eq_edit_sel")

        existing, eq_id = None, None
        if choice != "➕ เพิ่มใหม่":
            try:
                sel_code = choice.split(" — ")[0].strip()
                eq_data = df_all_eq[df_all_eq["code"] == sel_code]
                if not eq_data.empty:
                    existing = eq_data.iloc[0]
                    eq_id = int(existing["id"])
                else:
                    st.warning("⚠️ ไม่พบอุปกรณ์นี้ กรุณาเลือกใหม่")
            except Exception:
                st.warning("⚠️ เกิดข้อผิดพลาด กรุณาเลือกใหม่")

        if existing is not None:
            borrowed_now = int(existing["total_qty"]) - int(existing["available_qty"])
            st.markdown(
                f'<div class="eq-card" style="border-left:4px solid #1F4E79;">'
                f'📊 <b>ข้อมูลปัจจุบัน:</b> '
                f'ทั้งหมด <b>{existing["total_qty"]}</b> | '
                f'พร้อมใช้ <b>{existing["available_qty"]}</b> | '
                f'ยืมออก <b>{borrowed_now}</b>'
                f'</div>', unsafe_allow_html=True)

        form_key = f"frm_{eq_id if eq_id else 'new'}"
        with st.form(key=form_key):
            sv_code = st.text_input("รหัสอุปกรณ์ *",
                value=str(existing["code"]) if existing is not None else "")
            sv_name = st.text_input("ชื่ออุปกรณ์ *",
                value=str(existing["name"]) if existing is not None else "")
            sv_cat = st.text_input("หมวดหมู่",
                value=str(existing["category"] or "") if existing is not None else "")
            sv_qty = st.number_input("จำนวนทั้งหมด", min_value=1,
                value=int(existing["total_qty"]) if existing is not None else 1)
            sv_stat = st.selectbox("สถานะ", ["พร้อมใช้","ชำรุด","สูญหาย"],
                index=["พร้อมใช้","ชำรุด","สูญหาย"].index(str(existing["status"]))
                      if existing is not None else 0)
            sv_desc = st.text_area("รายละเอียด",
                value=str(existing["description"] or "") if existing is not None else "")
            sv_consumable = st.checkbox("📦 วัสดุสิ้นเปลือง (เบิกแล้วไม่ต้องคืน)",
                value=bool(existing.get("is_consumable", False)) if existing is not None else False)
            sv_min_qty = 0
            if sv_consumable:
                sv_min_qty = st.number_input("🔔 แจ้งเตือนเมื่อเหลือ ≤", min_value=0,
                    value=int(existing.get("min_qty", 5) or 5) if existing is not None else 5)
            sv_img = st.file_uploader(f"📷 รูปอุปกรณ์ (สูงสุด {MAX_UPLOAD_MB} MB)",
                                       type=["jpg","jpeg","png","gif"])

            if existing is not None and existing.get("image_url"):
                st.caption("รูปปัจจุบัน:")
                show_image(existing["image_url"], width="100%", size="preview")

            submitted = st.form_submit_button("💾 บันทึก", type="primary", use_container_width=True)

        if submitted:
            sv_code = sv_code.strip()
            sv_name = sv_name.strip()
            if not sv_code or not sv_name:
                st.error("กรุณากรอกรหัสและชื่ออุปกรณ์")
            else:
                dup = query_table("equipment", select="id", filters=[("code","eq",sv_code)])
                cur_id = eq_id if eq_id else -1
                if not dup.empty and (eq_id is None or int(dup.iloc[0]["id"]) != cur_id):
                    st.error(f"❌ รหัส '{sv_code}' มีอยู่แล้ว")
                else:
                    try:
                        img_url = existing.get("image_url") if existing is not None else None
                        if sv_img:
                            img_url = upload_image(sv_img, sv_code)
                            if img_url is None and sv_img:
                                st.stop()  # upload failed — อย่าบันทึก

                        if existing is None:
                            insert_row("equipment", {
                                "code": sv_code, "name": sv_name,
                                "category": sv_cat or None,
                                "total_qty": sv_qty, "available_qty": sv_qty,
                                "status": sv_stat, "image_url": img_url,
                                "description": sv_desc or None,
                                "is_consumable": sv_consumable,
                                "min_qty": sv_min_qty if sv_consumable else None
                            })
                            st.success(f"✅ บันทึกการเพิ่มอุปกรณ์ '{sv_code} — {sv_name}' เรียบร้อยแล้ว")
                            st.session_state["_next_sel"] = "➕ เพิ่มใหม่"
                        else:
                            old_total = int(existing["total_qty"])
                            old_available = int(existing["available_qty"])
                            borrowed = old_total - old_available
                            diff = int(sv_qty) - old_total
                            new_available = old_available + diff
                            if new_available < 0:
                                st.error(f"❌ ลดจำนวนไม่ได้! ยืมออกอยู่ {borrowed}")
                                st.stop()
                            update_rows("equipment", {
                                "code": sv_code, "name": sv_name,
                                "category": sv_cat or None,
                                "total_qty": sv_qty, "available_qty": new_available,
                                "status": sv_stat, "image_url": img_url,
                                "description": sv_desc or None,
                                "is_consumable": sv_consumable,
                                "min_qty": sv_min_qty if sv_consumable else None
                            }, "id", eq_id)
                            msg = f"✅ บันทึกการแก้ไข '{sv_code} — {sv_name}' เรียบร้อยแล้ว"
                            if diff > 0:
                                msg += f" (เพิ่มจำนวน +{diff} พร้อมใช้: {new_available})"
                            elif diff < 0:
                                msg += f" (ลดจำนวน {diff} พร้อมใช้: {new_available})"
                            st.success(msg)
                            st.session_state["_next_sel"] = f"{sv_code} — {sv_name}"
                        clear_all_cache()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ เกิดข้อผิดพลาด: {e}")
    else:
        st.info("🔒 การเพิ่ม/แก้ไขอุปกรณ์ สำหรับ Admin เท่านั้น")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: BORROW                                                  [FIX #2, #8]
# ═════════════════════════════════════════════════════════════════════════════
# ─── Duplicate Submit Guard ───────────────────────────────────────────────────
DUPLICATE_GUARD_SEC = 10

def _make_borrow_key(eq_id, name):
    return f"{eq_id}|{str(name).strip().lower()}"

def check_dup(eq_id, name):
    if st.session_state.get("lab_is_submitting"):
        return "submitting"
    last = st.session_state.get("lab_last_submit")
    if last and last["key"] == _make_borrow_key(eq_id, name):
        if time.time() - last["ts"] < DUPLICATE_GUARD_SEC:
            return "duplicate"
    return "ok"

def register_borrow(eq_id, name):
    st.session_state["lab_is_submitting"] = True
    st.session_state["lab_last_submit"] = {
        "key": _make_borrow_key(eq_id, name), "ts": time.time()
    }

def finish_borrow():
    st.session_state.pop("lab_is_submitting", None)

def clear_dup_guard():
    st.session_state.pop("lab_last_submit", None)
    st.session_state.pop("lab_is_submitting", None)

# ─── Step Wizard Helper ────────────────────────────────────────────────────────
def render_step_bar(current, steps):
    html = '<div class="step-bar">'
    for i, label in enumerate(steps):
        n = i + 1
        if n < current:
            cc, lc, ct = "done", "", "✓"
        elif n == current:
            cc, lc, ct = "active", "active", str(n)
        else:
            cc, lc, ct = "", "", str(n)
        if i > 0:
            lnc = "done" if i < current - 1 else ""
            html += f'<div class="step-line {lnc}"></div>'
        html += (f'<div class="step-item">'
                 f'<div class="step-circle {cc}">{ct}</div>'
                 f'<div class="step-label {lc}">{label}</div></div>')
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# PAGE: BORROW — 2-Step Wizard
# ═════════════════════════════════════════════════════════════════════════════
def page_borrow():
    st.markdown('<div class="page-title">➕ เบิกอุปกรณ์</div>', unsafe_allow_html=True)

    # ── Success Screen ──────────────────────────────────────────────────────
    if st.session_state.get("borrow_success"):
        info = st.session_state.pop("borrow_success")
        st.markdown(f"""
        <div style="min-height:75vh;display:flex;flex-direction:column;
            align-items:center;justify-content:center;text-align:center;padding:24px 20px;">
            <div style="font-size:5rem;margin-bottom:12px;animation:pop 0.4s ease;">🎉</div>
            <div style="font-size:1.6rem;font-weight:800;color:#5c2018;margin-bottom:8px;">
                บันทึกสำเร็จ!</div>
            <div style="background:#fdf2f0;border:2px solid #e8c4bc;border-radius:16px;
                padding:20px 24px;max-width:420px;width:100%;text-align:left;color:#5c2018;
                margin:16px 0 24px 0;">
                <div style="margin:6px 0;font-size:1.05rem;">{"📦" if info["consumable"] else "🔬"} <b>{info["eq_name"]}</b> x{info["qty"]}</div>
                <div style="margin:6px 0;font-size:1.05rem;">👤 <b>{info["name"]}</b> ({info["btype"]})</div>
                <div style="margin:6px 0;font-size:1.05rem;">📅 วันที่เบิก: {info["bdate"]}</div>
                {"" if info["consumable"] else f'<div style="margin:6px 0;font-size:1.05rem;">🗓️ กำหนดคืน: <b>{info["due"]}</b></div>'}
                {"<div style='margin:6px 0;font-size:0.9rem;color:#a85240;'>📦 วัสดุสิ้นเปลือง — ไม่ต้องคืน</div>" if info["consumable"] else ""}
            </div>
        </div>
        <style>
        @keyframes pop {{
            0%   {{ transform: scale(0.5); opacity:0; }}
            70%  {{ transform: scale(1.2); }}
            100% {{ transform: scale(1); opacity:1; }}
        }}
        </style>
        """, unsafe_allow_html=True)
        st.balloons()
        cl, cc2, cr = st.columns([1, 2, 1])
        with cc2:
            if st.button("➕ เบิกอุปกรณ์อีกครั้ง", type="primary",
                         use_container_width=True, key="borrow_again"):
                st.rerun()
            if st.button("🏠 กลับหน้าหลัก", use_container_width=True, key="borrow_home"):
                st.session_state.page = "หน้าหลัก"
                st.rerun()
        return

    # ── Init step ──────────────────────────────────────────────────────────
    if "borrow_step" not in st.session_state:
        st.session_state.borrow_step = 1

    render_step_bar(st.session_state.borrow_step, ["เลือกอุปกรณ์", "ข้อมูลผู้เบิก", "ยืนยัน"])

    # ── Load equipment (แสดงทุกรายการ ยกเว้น ชำรุด/สูญหาย) ────────────────
    all_eq = query_table("equipment",
                         select="id,code,name,category,available_qty,status,image_url,description,is_consumable",
                         filters=[],
                         order=[("category",{"desc":False}),("code",{"desc":False})])
    if not all_eq.empty:
        all_eq = all_eq[~all_eq["status"].isin(["ชำรุด","สูญหาย"])]

    if all_eq.empty:
        st.warning("⚠️ ไม่มีอุปกรณ์ในระบบ")
        return

    # ── STEP 1: เลือกอุปกรณ์ ───────────────────────────────────────────────
    if st.session_state.borrow_step == 1:
        cats = ["ทั้งหมด"] + sorted(all_eq["category"].dropna().unique().tolist())
        cat_sel = st.selectbox("📂 กรองหมวดหมู่", cats, key="bcat")
        search_eq = st.text_input("🔍 ค้นหาอุปกรณ์ที่ต้องการ",
                                   placeholder="ชื่ออุปกรณ์ หรือ รหัส", key="bsearch")

        filtered = all_eq.copy()
        if cat_sel != "ทั้งหมด":
            filtered = filtered[filtered["category"] == cat_sel]
        if search_eq:
            mask = (filtered["name"].str.contains(search_eq, case=False, na=False) |
                    filtered["code"].str.contains(search_eq, case=False, na=False))
            filtered = filtered[mask]

        if filtered.empty:
            st.info("ไม่พบอุปกรณ์ที่ค้นหา")
            return

        # dropdown — แสดงสถานะถ้าหมด
        opts = {}
        for _, r in filtered.iterrows():
            av = int(r["available_qty"])
            lbl = (f"{r['code']} — {r['name']}  (คงเหลือ {av})"
                   if av > 0 else f"{r['code']} — {r['name']}  ⛔ ยืมออกทั้งหมด")
            opts[lbl] = r["id"]

        sel_lbl = st.selectbox("เลือกอุปกรณ์ *", list(opts.keys()), key="beq")
        eq_row  = all_eq[all_eq["id"] == opts[sel_lbl]].iloc[0]
        av_qty  = int(eq_row["available_qty"])

        ci, cr2 = st.columns([1, 2])
        with ci:
            show_image(eq_row.get("image_url"), width="100%", size="preview")
        with cr2:
            st.markdown(f"**{eq_row['code']}** — {eq_row['name']}")
            st.markdown(f"หมวด: {eq_row['category'] or '-'}")
            if av_qty > 0:
                st.markdown(f"คงเหลือ: **{av_qty}**")
            else:
                st.markdown('<b style="color:#D62828;">⛔ ยืมออกทั้งหมด</b>',
                            unsafe_allow_html=True)
            if eq_row.get("is_consumable"):
                st.markdown(badge("วัสดุสิ้นเปลือง", "#17a2b8"), unsafe_allow_html=True)
            if eq_row.get("description"):
                st.caption(eq_row["description"])

        if av_qty > 0:
            qty = st.number_input("จำนวน *", min_value=1, max_value=av_qty, value=1, key="bqty")
            if st.button("ถัดไป → กรอกข้อมูลผู้เบิก", type="primary",
                         use_container_width=True, key="borrow_next"):
                st.session_state["borrow_eq_id"]  = int(opts[sel_lbl])
                st.session_state["borrow_eq_qty"] = qty
                st.session_state.borrow_step = 2
                st.rerun()
        else:
            st.button("⛔ ไม่สามารถเบิกได้ — ยืมออกทั้งหมด",
                      use_container_width=True, key="borrow_next", disabled=True)
            st.info("💡 กรุณารอจนกว่าจะมีการคืนอุปกรณ์")

    # ── STEP 2: ข้อมูลผู้เบิก ──────────────────────────────────────────────
    elif st.session_state.borrow_step == 2:
        sel_id  = st.session_state.get("borrow_eq_id")
        sel_qty = st.session_state.get("borrow_eq_qty", 1)
        eq_match = all_eq[all_eq["id"] == sel_id]
        if eq_match.empty:
            st.error("❌ ไม่พบอุปกรณ์ กรุณาเริ่มใหม่")
            st.session_state.borrow_step = 1
            st.rerun()

        eq_row2 = eq_match.iloc[0]
        is_cons = bool(eq_row2.get("is_consumable", False))

        # สรุปอุปกรณ์ที่เลือก
        st.markdown(
            f'<div class="eq-card" style="border-left:4px solid #7d3020;">'
            f'{"📦" if is_cons else "🔬"} <b>{eq_row2["code"]}</b> — {eq_row2["name"]} '
            f'<span style="background:#fdf2f0;color:#5c2018;border-radius:12px;'
            f'padding:2px 8px;font-size:0.82rem;font-weight:700;"> x{sel_qty}</span>'
            f'{"<span style=\"color:#17a2b8;font-size:0.82rem;\"> วัสดุสิ้นเปลือง</span>" if is_cons else ""}'
            f'</div>', unsafe_allow_html=True)

        st.markdown('<div class="section-header">👤 ข้อมูลผู้เบิก</div>', unsafe_allow_html=True)
        borrower_type = st.radio("ประเภท", ["นักศึกษา", "บุคลากร/อาจารย์"],
                                  horizontal=True, key="btype")
        borrower_name = st.text_input("ชื่อ-นามสกุล *", placeholder="กรอกชื่อ-นามสกุล", key="bname")
        phone = st.text_input("เบอร์โทรศัพท์ *", placeholder="เช่น 081-234-5678", key="bphone")
        student_id = st.text_input("รหัสนักศึกษา / รหัสพนักงาน (ถ้ามี)", key="bsid")
        department = st.text_input("ภาควิชา / หน่วยงาน (ถ้ามี)", key="bdept")

        st.markdown('<div class="section-header">📅 วันที่</div>', unsafe_allow_html=True)
        cd1, cd2 = st.columns(2)
        borrow_date = cd1.date_input("วันที่เบิก *", value=date.today(), key="bdate")
        if is_cons:
            due_date = None
            st.info("📦 วัสดุสิ้นเปลือง — ไม่ต้องกำหนดวันคืน")
        else:
            due_date = cd2.date_input("กำหนดคืน *", value=date.today(), key="bdue")
        condition_out = st.selectbox("สภาพขณะเบิก",
                                     ["ปกติ", "มีรอยขีดข่วน", "ชำรุดบางส่วน"], key="bcond")
        note = st.text_input("หมายเหตุ (ถ้ามี)", key="bnote")

        col_back, col_sub = st.columns([1, 2])
        with col_back:
            if st.button("← ย้อนกลับ", use_container_width=True, key="borrow_back"):
                st.session_state.borrow_step = 1
                st.rerun()
        with col_sub:
            if st.button("✅ ยืนยันการเบิก", type="primary",
                         use_container_width=True, key="borrow_submit"):
                if not borrower_name.strip():
                    st.error("❌ กรุณากรอกชื่อ-นามสกุล")
                elif not phone.strip():
                    st.error("❌ กรุณากรอกเบอร์โทรศัพท์")
                elif not is_cons and due_date and due_date < borrow_date:
                    st.error("❌ วันกำหนดคืนต้องไม่ก่อนวันที่เบิก")
                elif check_dup(sel_id, borrower_name) == "submitting":
                    st.warning("⏳ กำลังบันทึก กรุณารอสักครู่...")
                elif check_dup(sel_id, borrower_name) == "duplicate":
                    st.session_state["borrow_pending_confirm"] = True
                else:
                    _do_borrow(sel_id, sel_qty, eq_row2, borrower_name, borrower_type,
                               phone, student_id, department, borrow_date, due_date,
                               condition_out, note, is_cons)

        # Duplicate confirm dialog
        if st.session_state.get("borrow_pending_confirm"):
            st.warning(
                f"⚠️ **กดซ้ำภายใน {DUPLICATE_GUARD_SEC} วินาที!**\n\n"
                f"รายการล่าสุดอาจถูกบันทึกไปแล้ว\n"
                f"👤 **{borrower_name}** — **{eq_row2['name']}** x{sel_qty}\n\n"
                f"ต้องการเบิก **2 รอบ** จริงหรือไม่?"
            )
            cy2, cn2 = st.columns(2)
            with cy2:
                if st.button("✅ ใช่ เบิก 2 รอบ", type="primary",
                             use_container_width=True, key="borrow_confirm_yes"):
                    clear_dup_guard()
                    st.session_state.pop("borrow_pending_confirm", None)
                    _do_borrow(sel_id, sel_qty, eq_row2, borrower_name, borrower_type,
                               phone, student_id, department, borrow_date, due_date,
                               condition_out, note, is_cons)
            with cn2:
                if st.button("❌ ยกเลิก", use_container_width=True, key="borrow_confirm_no"):
                    st.session_state.pop("borrow_pending_confirm", None)
                    st.info("ℹ️ ยกเลิกแล้ว ไม่ถูกบันทึกซ้ำ")
                    st.rerun()


def _do_borrow(eq_id, qty, eq_row, borrower_name, borrower_type, phone,
               student_id, department, borrow_date, due_date, condition_out, note, is_cons):
    """Execute borrow transaction"""
    try:
        register_borrow(eq_id, borrower_name)
        existing_borr = query_table("borrowers", select="id",
                                    filters=[("phone","eq",phone.strip())])
        if not existing_borr.empty:
            borr_id = int(existing_borr.iloc[0]["id"])
            update_rows("borrowers", {
                "name": borrower_name.strip(), "type": borrower_type,
                "student_id": student_id or None, "department": department or None
            }, "id", borr_id)
        else:
            borr = insert_row("borrowers", {
                "name": borrower_name.strip(), "type": borrower_type,
                "student_id": student_id or None, "department": department or None,
                "phone": phone.strip()
            })
            borr_id = int(borr["id"])

        tx_status = "เบิกแล้ว" if is_cons else "ยืมอยู่"
        insert_row("transactions", {
            "equipment_id": eq_id, "borrower_id": borr_id,
            "qty": qty, "borrow_date": str(borrow_date),
            "due_date": str(due_date) if due_date else None,
            "condition_out": condition_out,
            "note": note or None, "status": tx_status
        })
        new_avail = int(eq_row["available_qty"]) - qty
        update_rows("equipment", {"available_qty": new_avail}, "id", eq_id)
        if new_avail <= 0:
            update_rows("equipment", {"status": "ยืมออก"}, "id", eq_id)
        clear_all_cache()
        finish_borrow()

        # Set success screen
        st.session_state["borrow_success"] = {
            "eq_name":   eq_row["name"],
            "qty":       qty,
            "name":      borrower_name.strip(),
            "btype":     borrower_type,
            "bdate":     str(borrow_date),
            "due":       str(due_date) if due_date else "-",
            "consumable": is_cons,
        }
        st.session_state.borrow_step = 1
        for k in ["borrow_eq_id","borrow_eq_qty"]:
            st.session_state.pop(k, None)
        st.rerun()
    except Exception as e:
        finish_borrow()
        st.error(f"❌ เกิดข้อผิดพลาด: {e}")



# ═════════════════════════════════════════════════════════════════════════════
# PAGE: RETURN                                                  [FIX #1, #2]
# ═════════════════════════════════════════════════════════════════════════════
def page_return():
    st.markdown('<div class="page-title">✅ คืนอุปกรณ์</div>', unsafe_allow_html=True)

    # ── Return Success Screen ───────────────────────────────────────────────
    if st.session_state.get("return_success"):
        info = st.session_state.pop("return_success")
        st.markdown(f"""
        <div style="min-height:75vh;display:flex;flex-direction:column;
            align-items:center;justify-content:center;text-align:center;padding:24px 20px;">
            <div style="font-size:5rem;margin-bottom:12px;animation:pop 0.4s ease;">📬</div>
            <div style="font-size:1.6rem;font-weight:800;color:#1a5276;margin-bottom:8px;">
                แจ้งคืนสำเร็จ!</div>
            <div style="font-size:0.95rem;color:#52796F;margin-bottom:16px;">
                รอ Admin ตรวจสอบและยืนยันรับคืน</div>
            <div style="background:#eaf4fb;border:2px solid #aed6f1;border-radius:16px;
                padding:20px 24px;max-width:420px;width:100%;text-align:left;color:#1a5276;
                margin:16px 0 24px 0;">
                <div style="margin:6px 0;font-size:1.05rem;">🔬 <b>{info["eq_name"]}</b> ({info["qty"]} ชิ้น)</div>
                <div style="margin:6px 0;font-size:1.05rem;">👤 {info["ret_name"]}</div>
                <div style="margin:6px 0;font-size:1.05rem;">📅 วันที่คืน: {info["ret_date"]}</div>
                <div style="margin:6px 0;font-size:1.05rem;">🔍 สภาพ: {info["cond"]}</div>
            </div>
        </div>
        <style>@keyframes pop {{
            0% {{ transform: scale(0.5); opacity:0; }}
            70% {{ transform: scale(1.2); }}
            100% {{ transform: scale(1); opacity:1; }}
        }}</style>
        """, unsafe_allow_html=True)
        cl, cc2, cr = st.columns([1, 2, 1])
        with cc2:
            if st.button("↩️ คืนอุปกรณ์อีกชิ้น", type="primary",
                         use_container_width=True, key="return_again"):
                st.rerun()
            if st.button("🏠 กลับหน้าหลัก", use_container_width=True, key="return_home"):
                st.session_state.page = "หน้าหลัก"
                st.rerun()
        return

    # นับจำนวนรอตรวจสอบ
    df_pending = load_pending_transactions_enriched()
    if not df_pending.empty and "id" in df_pending.columns:
        df_pending = df_pending.sort_values("id", ascending=False)
    n_pending = len(df_pending)
    tab2_label = f"🔍 รอตรวจสอบ ({n_pending})"

    tab1, tab2 = st.tabs(["📬 แจ้งคืนอุปกรณ์", tab2_label])

    # ── TAB 1: ผู้เบิกแจ้งคืน ────────────────────────────────────────────
    with tab1:
        st.markdown(
            '<div class="info-box">📬 ค้นหารายการที่ยืมอยู่ → กรอกข้อมูล → กด แจ้งคืน<br>'
            '<b>Admin จะตรวจสอบและยืนยันรับคืน</b></div>', unsafe_allow_html=True)
        search = st.text_input("🔍 พิมพ์ชื่อผู้ยืม หรือ รหัสอุปกรณ์",
                                placeholder="เช่น สมชาย / LAB-001", key="ret_search")

        df = load_active_transactions_enriched()

        if search and not df.empty:
            s = search.lower()
            mask = (df["br_name"].str.lower().str.contains(s, na=False) |
                    df["eq_code"].str.lower().str.contains(s, na=False) |
                    df["eq_name"].str.lower().str.contains(s, na=False))
            df = df[mask]

        if df.empty:
            st.info("✅ ไม่มีรายการยืม" if not search else "ไม่พบรายการที่ค้นหา")
        else:
            st.caption(f"พบ {len(df)} รายการ")
            for _, r in df.iterrows():
                od = overdue_days(r["due_date"])
                lbl_icon = "🔴" if od > 0 else "🟢"
                label = (f"{lbl_icon} **{r['eq_code']} — {r['eq_name']}** "
                         f"({r['qty']}) | 👤 {r['br_name']}"
                         + (f" ⚠️ เกิน {od} วัน" if od > 0 else ""))
                with st.expander(label, expanded=(od > 0)):
                    c1, c2 = st.columns([1, 2])
                    with c1:
                        show_image(r.get("eq_image_url"), width="100%", size="thumb")
                    with c2:
                        st.markdown(
                            f"📦 **{r['eq_code']}** — {r['eq_name']} ({r['qty']})<br>"
                            f"👤 {r['br_name']} ({r['br_type']})<br>"
                            f"📅 เบิก {r['borrow_date']} | คืน <b>{r['due_date']}</b>"
                            + (f"<br><b style='color:red;'>⚠️ เกิน {od} วัน</b>" if od > 0 else ""),
                            unsafe_allow_html=True)
                        st.caption(f"สภาพตอนเบิก: {r['condition_out']}")

                    col_rd, col_ci = st.columns(2)
                    return_date = col_rd.date_input("วันที่นำมาคืน",
                                                     value=date.today(), key=f"rd_{r['id']}")
                    condition_in = col_ci.selectbox("สภาพอุปกรณ์",
                                                     ["ปกติ", "มีรอยขีดข่วน", "ชำรุด", "สูญหาย"],
                                                     key=f"ci_{r['id']}")
                    return_note = st.text_input("หมายเหตุ (ถ้ามี)", key=f"rn_{r['id']}")

                    if st.button("📬 แจ้งคืน", key=f"notify_{r['id']}",
                                 type="primary", use_container_width=True):
                        borrow_dt = datetime.strptime(str(r["borrow_date"]), "%Y-%m-%d").date()
                        if return_date < borrow_dt:
                            st.error(f"❌ วันที่คืนต้องไม่ก่อนวันที่เบิก ({r['borrow_date']})")
                        else:
                            update_rows("transactions", {
                                "return_date": str(return_date), "condition_in": condition_in,
                                "note": return_note or None, "status": "รอตรวจสอบ"
                            }, "id", r["id"])
                            clear_all_cache()
                            st.session_state["return_success"] = {
                                "eq_name":  r["eq_name"],
                                "qty":      r["qty"],
                                "ret_name": r["br_name"],
                                "ret_date": str(return_date),
                                "cond":     condition_in,
                            }
                            st.rerun()

    # ── TAB 2: Admin ตรวจสอบ ─────────────────────────────────────────────
    with tab2:
        df_wait = df_pending

        if df_wait.empty:
            st.info("✅ ไม่มีรายการรอตรวจสอบ")
        else:
            st.warning(f"🔍 มี {len(df_wait)} รายการรอ Admin ตรวจสอบ")
            for _, r in df_wait.iterrows():
                od = overdue_days(r["due_date"])
                label = f"TX#{r['id']} | {r['eq_code']} {r['eq_name']} | {r['br_name']}"

                with st.expander(label):
                    c1, c2 = st.columns([1, 2])
                    with c1:
                        show_image(r.get("eq_image_url"), width="100%", size="preview")
                    with c2:
                        phone_str = f"📞 {r['br_phone']}" if pd.notna(r.get('br_phone')) and r.get('br_phone') else ""
                        st.markdown(
                            f"📦 **{r['eq_code']}** — {r['eq_name']} ({r['qty']})<br>"
                            f"👤 {r['br_name']} ({r['br_type']}) {phone_str}<br>"
                            f"📅 เบิก {r['borrow_date']} | กำหนดคืน <b>{r['due_date']}</b><br>"
                            f"📅 แจ้งคืนวันที่: <b>{r.get('return_date','')}</b>"
                            + (f"<br><b style='color:red;'>⚠️ เกินกำหนด {od} วัน</b>" if od > 0 else ""),
                            unsafe_allow_html=True)
                        st.markdown(f"**สภาพตอนเบิก:** {r['condition_out']}")
                        st.markdown(f"**สภาพที่แจ้งคืน:** {r.get('condition_in') or '-'}")
                        if r.get("note"):
                            st.caption(f"หมายเหตุ: {r['note']}")

                    if is_admin():
                        st.markdown("**Admin ตรวจสอบจริง:**")
                        admin_condition = st.selectbox("สภาพอุปกรณ์จริงที่ตรวจสอบ",
                                                       ["ปกติ", "มีรอยขีดข่วน", "ชำรุด", "สูญหาย"],
                                                       key=f"ac_{r['id']}")
                        admin_note = st.text_input("หมายเหตุ Admin", key=f"an_{r['id']}")

                        col_ok, col_rej = st.columns(2)
                        with col_ok:
                            if st.button("✅ ยืนยันรับคืน", key=f"confirm_{r['id']}",
                                         type="primary", use_container_width=True):
                                note_final = f"[Admin: {admin_note}]" if admin_note else r.get("note")
                                update_rows("transactions", {
                                    "condition_in": admin_condition, "note": note_final,
                                    "status": "คืนแล้ว"
                                }, "id", r["id"])

                                cur_eq = query_table("equipment", select="available_qty",
                                                     filters=[("id","eq",r["equipment_id"])])
                                if not cur_eq.empty:
                                    cur_avail = int(cur_eq.iloc[0]["available_qty"])
                                    qty_returned = int(r["qty"])

                                    # Fix 8: logic ชัดเจน — สูญหาย = ไม่ได้คืนจริง
                                    if admin_condition == "สูญหาย":
                                        available_delta = 0  # ไม่เพิ่ม available
                                        new_status = "สูญหาย"
                                    elif admin_condition == "ชำรุด":
                                        available_delta = qty_returned  # คืนแต่ชำรุด
                                        new_status = "ชำรุด"
                                    else:
                                        available_delta = qty_returned  # คืนปกติ
                                        new_status = "พร้อมใช้"

                                    update_rows("equipment", {
                                        "available_qty": cur_avail + available_delta,
                                        "status": new_status
                                    }, "id", r["equipment_id"])

                                clear_all_cache()
                                st.success(f"✅ ยืนยันรับคืนแล้ว สภาพ: {admin_condition}")
                                st.rerun()
                        with col_rej:
                            if st.button("↩️ ส่งกลับ (ยังไม่คืน)", key=f"reject_{r['id']}",
                                         use_container_width=True):
                                update_rows("transactions", {
                                    "status": "ยืมอยู่", "return_date": None, "condition_in": None
                                }, "id", r["id"])
                                clear_all_cache()
                                st.warning("↩️ ส่งกลับเป็นสถานะ ยืมอยู่ แล้ว")
                                st.rerun()
                    else:
                        st.info("🔒 กรุณา Login Admin เพื่อยืนยันรับคืน")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: REPORT                                                  [FIX #1]
# ═════════════════════════════════════════════════════════════════════════════
def page_report():
    st.markdown('<div class="page-title">📋 รายงาน</div>', unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["ประวัติการเบิก-คืน", "สรุปอุปกรณ์"])

    with tab1:
        date_from = st.date_input("ตั้งแต่", value=date(date.today().year, 1, 1))
        date_to = st.date_input("ถึงวันที่", value=date.today())
        status_filter = st.selectbox("สถานะ", ["ทั้งหมด", "ยืมอยู่", "คืนแล้ว", "เบิกแล้ว"])

        # [FIX #1] batch fetch + merge
        df_report = load_report_data(date_from, date_to, status_filter)

        st.caption(f"พบ {len(df_report)} รายการ")
        st.dataframe(df_report, use_container_width=True, hide_index=True)

        if not df_report.empty:
            if is_admin():
                st.download_button("📥 Export Excel", data=export_excel(df_report, "ประวัติการเบิก-คืน"),
                                   file_name=f"borrow_history_{date.today()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
            else:
                st.info("🔒 Export Excel สำหรับ Admin เท่านั้น")

    with tab2:
        # [FIX #1] batch fetch + groupby
        df2 = load_equipment_summary()

        if not df2.empty:
            st.dataframe(df2, use_container_width=True, hide_index=True)
            if is_admin():
                st.download_button("📥 Export Excel สรุปอุปกรณ์",
                                   data=export_excel(df2, "สรุปอุปกรณ์"),
                                   file_name=f"equipment_summary_{date.today()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
        else:
            st.info("ยังไม่มีข้อมูลอุปกรณ์")

# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def export_excel(df, sheet_name):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill  = PatternFill("solid", fgColor="5c2018")
    hfont  = Font(color="FFFFFF", bold=True, size=11)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill, cell.font, cell.border = hfill, hfont, border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="FDF2F0" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border, cell.fill = border, fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 30)
    ws.row_dimensions[1].height = 25
    wb.save(output)
    return output.getvalue()

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: SETTINGS                                                [FIX #4, #7]
# ═════════════════════════════════════════════════════════════════════════════
def page_settings():
    st.markdown('<div class="page-title">⚙️ ตั้งค่าระบบ</div>', unsafe_allow_html=True)

    if not is_admin():
        st.warning("🔒 หน้านี้สำหรับ Admin เท่านั้น กรุณา Login ที่ Sidebar")
        return

    tab1, tab2, tab3 = st.tabs(["💾 สำรองข้อมูล", "📂 นำเข้าข้อมูล", "🗑️ ล้างข้อมูล"])

    with tab1:
        st.markdown('<div class="section-header">💾 Export — สำรองข้อมูลเป็น JSON</div>', unsafe_allow_html=True)
        st.info("Export ข้อมูลทั้งหมดเป็นไฟล์ JSON สำหรับสำรองหรือย้ายระบบ")

        if st.button("📦 สร้างไฟล์ Backup JSON", type="primary", use_container_width=True):
            eq = query_table("equipment").to_dict(orient="records")
            tx = query_table("transactions").to_dict(orient="records")
            borr = query_table("borrowers").to_dict(orient="records")
            backup = {
                "backup_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "version": "2.1-optimized",
                "equipment": eq, "transactions": tx, "borrowers": borr
            }
            json_str = json.dumps(backup, ensure_ascii=False, indent=2, default=str)
            fname = f"lab_backup_{date.today()}.json"
            st.download_button(
                label=f"⬇️ ดาวน์โหลด {fname}",
                data=json_str.encode("utf-8"),
                file_name=fname, mime="application/json",
                use_container_width=True
            )
            st.success(f"✅ สร้าง Backup สำเร็จ — อุปกรณ์ {len(eq)} | ประวัติ {len(tx)}")

    with tab2:
        st.markdown('<div class="section-header">📂 Import — นำเข้าข้อมูลจาก JSON</div>', unsafe_allow_html=True)
        st.warning("⚠️ การนำเข้าจะ **เพิ่ม** ข้อมูลเข้าระบบ ไม่ได้ลบข้อมูลเดิม")

        uploaded_json = st.file_uploader("เลือกไฟล์ JSON", type=["json"], key="import_json")
        if uploaded_json:
            try:
                data = json.loads(uploaded_json.read().decode("utf-8"))
                eq_count = len(data.get("equipment", []))
                tx_count = len(data.get("transactions", []))
                borr_count = len(data.get("borrowers", []))
                backup_date = data.get("backup_date", "ไม่ทราบ")

                st.info(f"📋 Backup วันที่: {backup_date} | อุปกรณ์: {eq_count} | ประวัติ: {tx_count} | ผู้เบิก: {borr_count}")

                import_mode = st.radio("โหมดนำเข้า",
                    ["เฉพาะอุปกรณ์ (equipment)", "ทั้งหมด (equipment + transactions + borrowers)"])

                if st.button("📂 ยืนยันนำเข้าข้อมูล", type="primary", use_container_width=True):
                    imported = 0
                    for eq in data.get("equipment", []):
                        try:
                            existing = query_table("equipment", select="id",
                                                   filters=[("code","eq",eq["code"])])
                            if existing.empty:
                                img_url = eq.get("image_url") or eq.get("image_path")
                                if img_url and not img_url.startswith("http"):
                                    img_url = None
                                insert_row("equipment", {
                                    "code": eq["code"], "name": eq["name"],
                                    "category": eq.get("category"),
                                    "total_qty": eq.get("total_qty", 1),
                                    "available_qty": eq.get("available_qty", 1),
                                    "status": eq.get("status", "พร้อมใช้"),
                                    "image_url": img_url,
                                    "description": eq.get("description")
                                })
                                imported += 1
                        except Exception:
                            pass

                    if "ทั้งหมด" in import_mode:
                        id_map = {}
                        for b in data.get("borrowers", []):
                            old_id = b["id"]
                            result = insert_row("borrowers", {
                                "name": b["name"], "type": b["type"],
                                "student_id": b.get("student_id"),
                                "department": b.get("department"),
                                "phone": b.get("phone")
                            })
                            if result:
                                id_map[old_id] = result["id"]

                        for tx in data.get("transactions", []):
                            try:
                                new_borr_id = id_map.get(tx["borrower_id"])
                                eq_id = tx.get("equipment_id")
                                if new_borr_id and eq_id:
                                    insert_row("transactions", {
                                        "equipment_id": eq_id, "borrower_id": new_borr_id,
                                        "qty": tx.get("qty", 1),
                                        "borrow_date": tx.get("borrow_date"),
                                        "due_date": tx.get("due_date"),
                                        "return_date": tx.get("return_date"),
                                        "condition_out": tx.get("condition_out", "ปกติ"),
                                        "condition_in": tx.get("condition_in"),
                                        "note": tx.get("note"),
                                        "status": tx.get("status", "คืนแล้ว")
                                    })
                            except Exception:
                                pass

                    clear_all_cache()
                    st.success("✅ นำเข้าข้อมูลสำเร็จ!")
                    st.rerun()

            except Exception as e:
                st.error(f"❌ ไฟล์ไม่ถูกต้อง: {e}")

    # ── TAB 3: ล้างข้อมูล ─────────────────────────────────────────────────
    with tab3:
        st.markdown('<div class="section-header">🗑️ ล้างข้อมูล</div>', unsafe_allow_html=True)
        st.error("⚠️ การล้างข้อมูลไม่สามารถกู้คืนได้ แนะนำให้ **Export JSON ก่อน** ทุกครั้ง!")

        clear_mode = st.selectbox("เลือกประเภทการล้าง", [
            "เลือก...",
            "🔄 รีเซ็ตจำนวนอุปกรณ์ (available = total)",
            "📋 ล้างประวัติการเบิก-คืนทั้งหมด",
            "💥 ล้างทุกอย่าง (เริ่มระบบใหม่)",
        ])

        if clear_mode != "เลือก...":
            desc = {
                "🔄 รีเซ็ตจำนวนอุปกรณ์ (available = total)": "รีเซ็ต available_qty = total_qty ทุกอุปกรณ์ เปลี่ยนสถานะเป็น พร้อมใช้",
                "📋 ล้างประวัติการเบิก-คืนทั้งหมด":          "ลบ transactions + borrowers ข้อมูลอุปกรณ์ยังอยู่",
                "💥 ล้างทุกอย่าง (เริ่มระบบใหม่)":           "ลบทุกตาราง เริ่มใหม่ทั้งหมด",
            }
            st.info(f"ℹ️ {desc.get(clear_mode,'')}")

            st.markdown("**พิมพ์ CONFIRM เพื่อยืนยัน:**")
            confirm_text = st.text_input("", placeholder="พิมพ์ CONFIRM", key="confirm_clear")

            if st.button("🗑️ ดำเนินการล้างข้อมูล", type="primary", use_container_width=True):
                if confirm_text != "CONFIRM":
                    st.error("❌ กรุณาพิมพ์ CONFIRM ให้ถูกต้อง (ตัวพิมพ์ใหญ่)")
                else:
                    try:
                        if "รีเซ็ตจำนวน" in clear_mode:
                            # [FIX #4] batch reset
                            batch_reset_equipment()
                            st.success("✅ รีเซ็ตจำนวนอุปกรณ์เรียบร้อย")

                        elif "ล้างประวัติ" in clear_mode:
                            # [FIX #7] delete all
                            delete_rows("transactions", delete_all=True)
                            delete_rows("borrowers", delete_all=True)
                            batch_reset_equipment()
                            st.success("✅ ล้างประวัติการเบิก-คืนเรียบร้อย")

                        elif "ล้างทุกอย่าง" in clear_mode:
                            delete_rows("transactions", delete_all=True)
                            delete_rows("borrowers", delete_all=True)
                            delete_rows("equipment", delete_all=True)
                            st.success("✅ ล้างข้อมูลทั้งหมดเรียบร้อย เริ่มระบบใหม่ได้เลย")

                        clear_all_cache()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ เกิดข้อผิดพลาด: {e}")

# ─── FOOTER ───────────────────────────────────────────────────────────────────
def footer():
    st.markdown("---")
    st.markdown("""
    <div style="text-align:center; color:#888; font-size:0.82rem; padding:8px 0 16px 0; line-height:1.8;">
        🔬 ระบบบริหารจัดการเบิก-คืนอุปกรณ์ห้องปฏิบัติการ TTC<br>
        <b style="color:#4CAF50;">☁️ Cloud Edition v2.1</b> — Supabase + Cloudinary<br>
        พัฒนาโดย <b style="color:#1F4E79;">รศ.ดร.อิทธิพล มีผล</b><br>
        ภาควิชาครุศาสตร์โยธา &nbsp;|&nbsp; คณะครุศาสตร์อุตสาหกรรม<br>
        มหาวิทยาลัยเทคโนโลยีพระจอมเกล้าพระนครเหนือ (KMUTNB)
    </div>
    """, unsafe_allow_html=True)

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    nav()
    page = st.session_state.get("page", "หน้าหลัก")

    ADMIN_ONLY = {"Dashboard", "อุปกรณ์", "รายงาน", "ตั้งค่า"}
    if page in ADMIN_ONLY and not is_admin():
        st.warning("🔒 หน้านี้สำหรับ Admin เท่านั้น กรุณา Login ที่เมนู ☰ มุมซ้ายบน")
        st.session_state.page = "หน้าหลัก"
        st.rerun()

    if   page == "หน้าหลัก": page_home()
    elif page == "Dashboard": page_dashboard()
    elif page == "อุปกรณ์":  page_equipment()
    elif page == "เบิก":      page_borrow()
    elif page == "คืน":       page_return()
    elif page == "รายงาน":    page_report()
    elif page == "ตั้งค่า":   page_settings()

    st.markdown("---")
    st.markdown('''
    <div style="text-align:center;color:#999;font-size:0.8rem;padding:8px 0 16px 0;line-height:1.9;">
        🔬 ระบบบริหารจัดการอุปกรณ์ห้องปฏิบัติการ<br>
        <b style="color:#7d3020;">☁️ Cloud Edition</b> — Supabase + Cloudinary<br>
        พัฒนาโดย <b style="color:#5c2018;">รศ.ดร.อิทธิพล มีผล</b><br>
        ภาควิชาครุศาสตร์โยธา &nbsp;|&nbsp; คณะครุศาสตร์อุตสาหกรรม<br>
        มหาวิทยาลัยเทคโนโลยีพระจอมเกล้าพระนครเหนือ (KMUTNB)
    </div>
    ''', unsafe_allow_html=True)


if __name__ == "__main__":
    main()
